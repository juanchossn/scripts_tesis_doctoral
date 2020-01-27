#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jan  8 11:38:00 2019

@author: gossn
"""

#%% MODULES
import numpy as np
import pandas as pd
import openpyxl 
from openpyxl import load_workbook
import sys
import os
import fileinput
import glob
path0 = '/home/gossn/Dropbox/Documents/inSitu/Database'
sys.path.append(path0 + '/scripts')
import campaignProcessing as cp0

#%% CHOOSE CAMPAIGN(S)!!
campaigns = ['all']
## POSSIBLE INPUTS
# all
# campaign, list of campaigns
# region, list of regions

campaigns = cp0.campaignList(campaigns,path0)

print(campaigns)
#%% PROCESSORS!!

#campaignsASD  = ['RdP_20130430_PuntaPiedras', 'RdP_20131120_FIPCA', 'RdP_20131220_FIPCA', 'RdP_20140106_FIPCA', 'RdP_20140227_FIPCA', 'RdP_20140318_Muelle', 'RdP_20140411_FIPCA', 'RdP_20140415_Muelle', 'RdP_20140501_FIPCA', 'RdP_20140516_Muelle', 'RdP_20140619_FIPCA', 'RdP_20140819_FIPCA', 'RdP_20141122_Samborombon', 'RdP_20141217_Samborombon', 'RdP_20150423_Muelle', 'RdP_20151118_Muelle_TURBINET', 'RdP_20160421_Pleiades', 'RdP_20160925_Mancha', 'RdP_20170106_Muelle', 'RdP_20180404_Muelle_HYPERMAQ', 'RdP_20181105_Muelle_PLATAGUS', 'RdP_20181204_Houssay', 'Tagus_20180828', 'Tagus_20190617']
#
#campaignsTriOS = ['RdP_20160421_Pleiades', 'RdP_20160925_Mancha', 'RdP_20170106_Muelle', 'RdP_20180404_Muelle_HYPERMAQ', 'RdP_20181105_Muelle_PLATAGUS', 'RdP_20181204_Houssay', 'Tagus_20180828', 'Tagus_20190617']

for camp in campaigns:
    print('Processing campaign: ' + camp)
    # ASD
    cp0.asdProcess(camp,path0)
    # TRIOS
    cp0.triosProcess(camp,path0)
    # PIC2STATIONPIC
#    cp0.pic2stationPic(camp,path0)
    # CS CAMPBELL
#    cp0.campbellContinuous2Stations(camp,path0)
    # T Dogliotti & SPM Nechad
#    cp0.TDogliottiSPMNechad(camp,path0)
    # CHL Gitelson 2008 & RDI Shen 2019
#    cp0.ThreeBandModels(camp,path0)
    
#%% Overpasses

pathRegions   = path0 + '/regions'
pathXlsx      = path0 + '/general/overpasses'


radiometers  = ['ASD','Trios']
sensors      = ['Rhow','MA','MT','ME','MSI','OLI','OLCI','PHR1B','VIIRS']

overpasses = pd.DataFrame()

flagInit = True

for camp in campaigns:

    region = camp.split('_')[0]
    month  = camp.split('_')[1]
    pathCampaign  = pathRegions + '/' + region + '/' +  camp
    campaign = region + '_' + month

    stationInfo = pd.read_excel(pathCampaign + '/' + region + '_' + month + '.xlsx', sheet_name='stationInfo', skiprows=1)
    stationInfo['StationID'] = campaign + '_' + stationInfo['StationID'].astype(str)
    stationInfo = stationInfo.set_index('StationID')

    overpassCond = ~stationInfo['Overpasses'].isnull()
    stationInfo  = stationInfo.loc[overpassCond,:]

    overpasses = pd.concat([overpasses,stationInfo],axis=0)

overpasses.index.name = 'StationID'
wb = openpyxl.Workbook()
wb.save(pathXlsx + '.xlsx')
writer = pd.ExcelWriter(pathXlsx + '.xlsx', engine = 'openpyxl')
writer.book = wb
del wb['Sheet']

# Write to Excel: StationInfo
sheetname = 'stationInfo'
if sheetname in wb.sheetnames:
    del wb[sheetname]
overpasses.to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)

cp0.adjustColWidth(wb.get_sheet_by_name(sheetname))
writer.save()
writer.close()

#%%  SCALARS!

pathRegions   = path0 + '/regions'
pathXlsx      = path0 + '/general' + '/scalars.xlsx'

#Tagus_20180828 RdP_20121113_Muelle_SeaSWIR
scalars = pd.DataFrame()
for camp in campaigns:
    print('Campaign: ' + camp)
    region = camp.split('_')[0]
    month  = camp.split('_')[1]
    pathCampaign  = pathRegions + '/' + region + '/' +  camp
    campaign = region + '_' + month


    M = {'T_HACH[FNU]'      :{'xlsxFileName': pathCampaign + '/' + region + '_' + month + '.xlsx'                           ,\
                      'sheetName'   : {'Mean': 'turbidityHACH'        , 'CV': 'turbidityHACH'          }             ,\
                      'header'      : {'Mean': ('globalMean[FNU]','') , 'CV': ('globalCV[%]','')       }             ,\
                      'skiprows'    : 0                                                                             },\

        'T_Dogliotti[FNU]' :{'xlsxFileName': pathCampaign + '/' + region + '_' + month + '_TDogliotti.xlsx'                 ,\
                      'sheetName'   : {'Mean': 'TDogliotti'           , 'CV': 'TDogliotti'             }             ,\
                      'header'      : {'Mean': 'Hyper'                , 'CV': 'Hyper_CV'               }             ,\
                      'skiprows'    : 1                                                                             },\

        'T_Dog_w[-]'       :{'xlsxFileName': pathCampaign + '/' + region + '_' + month + '_TDogliotti.xlsx'                 ,\
                      'sheetName'   : {'Mean': 'TDogliotti'           , 'CV': 'TDogliotti'             }             ,\
                      'header'      : {'Mean': 'Weight[0:Red,1:NIR]' , 'CV': 'Weight[0:Red,1:NIR]_CV'  }             ,\
                      'skiprows'    : 1                                                                             },\

        'SPM[mg/l]'       :{'xlsxFileName': pathCampaign + '/' + region + '_' + month + '.xlsx'                             ,\
                      'sheetName'   : {'Mean': 'SPM'                  , 'CV': 'SPM'                    }             ,\
                      'header'      : {'Mean': ('SPM[g/m3]','Mean')   , 'CV': ('SPM[g/m3]','CV[%]')    }             ,\
                      'skiprows'    : 0                                                                             },\

        'SOM[mg/l]'       :{'xlsxFileName': pathCampaign + '/' + region + '_' + month + '.xlsx'                             ,\
                      'sheetName'   : {'Mean': 'SPM'                  , 'CV': 'SPM'                    }             ,\
                      'header'      : {'Mean': ('SOM[g/m3]','Mean')   , 'CV': ('SOM[g/m3]','CV[%]')    }             ,\
                      'skiprows'    : 0                                                                             },\

        'SPM_Nechad[mg/l]' :{'xlsxFileName': pathCampaign + '/' + region + '_' + month + '_SPMNechad.xlsx'                  ,\
                      'sheetName'   : {'Mean': 'SPMNechad'            , 'CV': 'SPMNechad'              }             ,\
                      'header'      : {'Mean': 'Hyper'                , 'CV': 'Hyper_CV'               }             ,\
                      'skiprows'    : 1                                                                             },\

        'T_OBS501_BS[FBU]'  :{'xlsxFileName': pathCampaign + '/campbellProcessed/' + region + '_' + month + '_Campbell.xlsx',\
                      'sheetName'   : {'Mean': 'Stations_Mean'        , 'CV': 'Stations_CV'            }             ,\
                      'header'      : {'Mean': 'BS_OBS501_Global[FBU]', 'CV': 'BS_OBS501_Global[FBU]'  }             ,\
                      'skiprows'    : 1                                                                             },\

        'T_OBS501_SS[FNU]'  :{'xlsxFileName': pathCampaign + '/campbellProcessed/' + region + '_' + month + '_Campbell.xlsx',\
                      'sheetName'   : {'Mean': 'Stations_Mean'        , 'CV': 'Stations_CV'            }             ,\
                      'header'      : {'Mean': 'SS_OBS501_Global[FNU]', 'CV': 'SS_OBS501_Global[FNU]'  }             ,\
                      'skiprows'    : 1                                                                             },\

        'CHL[ug/l]'       :{'xlsxFileName': pathCampaign + '/' + region + '_' + month + '.xlsx'                             ,\
                      'sheetName'   : {'Mean': 'CHL'                  , 'CV': 'CHL'                    }             ,\
                      'header'      : {'Mean': ('CHL[ug/l]','mean')   , 'CV': ('CHL[ug/l]','CV[%]')    }             ,\
                      'skiprows'    : 0                                                                             },\
        'CHL_Gitelson2008[-]' :{'xlsxFileName': pathCampaign + '/' + region + '_' + month + '_CHL_Gitelson2008.xlsx'        ,\
                      'sheetName'   : {'Mean': 'CHL_Gitelson2008'           , 'CV': 'CHL_Gitelson2008' }             ,\
                      'header'      : {'Mean': 'Hyper'                , 'CV': 'Hyper_CV'               }             ,\
                      'skiprows'    : 1                                                                             },\
        'RDI_Shen2019[-]' :{'xlsxFileName': pathCampaign + '/' + region + '_' + month + '_RDI_Shen2019.xlsx'                ,\
                      'sheetName'   : {'Mean': 'RDI_Shen2019'           , 'CV': 'RDI_Shen2019'         }             ,\
                      'header'      : {'Mean': 'Hyper'                , 'CV': 'Hyper_CV'               }             ,\
                      'skiprows'    : 1                                                                             },\
       }


    scalars0 = pd.read_excel(pathCampaign + '/' + region + '_' + month + '.xlsx', sheet_name='stationInfo', skiprows=1)
    scalars0['StationID'] = campaign + '_' + scalars0['StationID'].astype(str)
    scalars0 = scalars0.set_index('StationID')

    scalars0 = scalars0.drop(['Notes','Overpasses'], axis=1)
    
    for m in M.keys():
        for stat in ['Mean','CV']:
#            del scalar
            try:
                if isinstance(M[m]['header'][stat], tuple):
                    headerLines = list(range(len(M[m]['header'][stat])))
                else:
                    headerLines = [0]

                scalar = pd.read_excel(M[m]['xlsxFileName'],header=headerLines, sheet_name=M[m]['sheetName'][stat],skiprows=M[m]['skiprows'])

                headerRedefUnnamed = []
                for head in headerLines:
                    headerRedef = list(scalar.columns.get_level_values(level=head).str.replace('Un.*',''))
                    headerRedefUnnamed.append(headerRedef)
                if headerLines == [0]:
                    headerRedefUnnamed = headerRedefUnnamed[0]
                scalar.columns = headerRedefUnnamed
                
                                # Solving BUG in pandas (if first station of a 'multiheader' sheet is empty, it doesn't take it as part of the stations)
                if np.shape(scalar)[0] != np.shape(scalars0)[0]:
                    data = pd.DataFrame(columns=[scalar.columns[j] for j in range(np.shape(scalar)[1])])
                    data.loc[scalar.index.name,:] = np.nan
                    scalar = pd.concat([data,scalar])
                    scalar.index.name = 'StationID'
#                    data = []
#                    cols = [[scalar.columns[j],np.nan] for j in range(np.shape(scalar)[1])]
#                    data.insert(0, cols)
#                    scalar = pd.concat([pd.Series([np.nan],index=[scalar.index.name]), scalar])
#                    pd.concat([scalar, ], ignore_index=True)
#                    scalar = pd.concat([pd.Series([np.nan],index=[scalar.index.name]), scalar])

                # Change IDs and name to append to a general DF
                scalar.index = scalars0.index.tolist()
                scalar = scalar[M[m]['header'][stat]]
                scalar.name  = m + '_' + stat
            except:
                if stat == 'Mean':
                    print('No ' + m + ' measurements for this campaign!')
                scalar = pd.DataFrame(columns = [m + '_' + stat], index=scalars0.index)
                
            scalars0 = pd.concat([scalars0,scalar],axis=1)
            continue
    
    scalars = pd.concat([scalars,scalars0],axis=0)

scalars.index.name = 'StationID'

wb = openpyxl.Workbook()
wb.save(pathXlsx)
writer = pd.ExcelWriter(pathXlsx, engine = 'openpyxl')
writer.book = wb
del wb['Sheet']

# Write to Excel: scalars0
sheetname = 'scalars'
if sheetname in wb.sheetnames:
    del wb[sheetname]
scalars.to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)

writer.save()
writer.close()

#%% Compiled ASD & TRIOS (OJO: ES necesario correr la seccion anterior!!)
radiometers  = ['ASD','Trios']
sensors      = ['Rhow','MA','MT','ME','MSI','OLI','OLCI','PHR1B','VIIRS']

for stl in ['overpasses','all']:
    if stl == 'overpasses':
        pathStl = path0 + '/general/overpasses.xlsx'
        sheetname = 'stationInfo'
    elif stl == 'all':
        pathStl = path0 + '/general/scalars.xlsx'
        sheetname = 'scalars'
    pathXlsx = path0 + '/general/' + stl
    stList = pd.read_excel(pathStl, sheet_name=sheetname, skiprows=1).set_index('StationID')

    flag = {}
    rhow = {}
    for rad in radiometers:
        flag[rad] = {}
        rhow[rad] = {}
        for sensor in sensors:
            for stat in ['','Std']:
                flag[rad][sensor + stat] = True
    
    for st in stList.index:
        print(st)
        [region,month] = st.split('_')[:2]
        st0 = '_'.join(st.split('_')[2:])
        camp = region + '_' + month
        pathCampaign  = pathRegions + '/' + region + '/' + camp
        for rad in radiometers:
            try:
                pathRhows = glob.glob(pathCampaign + '*/' + rad + 'Processed/' + camp + '*SatSensors.xlsx')[0]
            except:
                print('No ' + rad + ' measurements')
                continue
            for sensor in sensors:
                for stat in ['','Std']:
                    data = pd.read_excel(pathRhows, sheet_name=sensor + stat, skiprows=1)
                    if rad == 'ASD' and sensor == 'Rhow':
                        data  = data.T
                        data.columns = data.iloc[0]
                        data  = data.iloc[1:,:]
                        data.index.name = 'StationID'
    #                    new_index = [campaign + '_' + st for st in data.index]
    #                    data = data.reindex(new_index)
                    else:
                        data = data.rename(columns={'Unnamed: 0':'StationID'})
                        data = data.set_index('StationID')
                        data.index.name = 'StationID'
    #                    data['StationID'] = campaign + '_' + data['StationID'].astype(str)
    #                data = data.loc[overpassCond,:]
    ##                if len(rhow[rad][sensor + stat].columns) == 0:
    ##                    rhow[rad][sensor + stat] = pd.concat([rhow[rad][sensor + stat],pd.DataFrame(columns=data.columns)],axis=1)
                    if flag[rad][sensor + stat]:
                        rhow[rad][sensor + stat] = pd.DataFrame(index=stList.index,columns=data.columns)
                        flag[rad][sensor + stat] = False
                    rhow[rad][sensor + stat].loc[st,:] = data.loc[st0,:]
    #rhow[rad][sensor + stat].index.name = 'StationID'
    ##########################
    for rad in radiometers:
        wb = openpyxl.Workbook()
        wb.save(pathXlsx + '_' + rad + '.xlsx')
        writer = pd.ExcelWriter(pathXlsx + '_' + rad + '.xlsx', engine = 'openpyxl')
        writer.book = wb
        del wb['Sheet']

        for sensor in sensors:
            for stat in ['','Std']:
                # Write to Excel: StationInfo
                sheetname = sensor + stat
                if sheetname in wb.sheetnames:
                    del wb[sheetname]
                if rad == 'ASD' and sensor == 'Rhow':
                    rhow[rad][sensor + stat].T.to_excel(writer, index = True, sheet_name=sheetname,startrow=1)
                else:
                    rhow[rad][sensor + stat].to_excel(writer, index = True, sheet_name=sheetname,startrow=1)
                cp0.adjustColWidth(wb.get_sheet_by_name(sheetname))
                writer.save(); writer.save()
                writer.close(); writer.close()
#%% Instruments per campaign

pathRegions   = path0 + '/regions'
pathXlsx      = path0 + '/general/campaignsInventory.xlsx'


M = {'ASD'         :{'file':'rad' ,'sheet':'MA'           ,'skiprows':1,'header':0    ,'mag':412                    }                  ,\
     'Trios'       :{'file':'rad' ,'sheet':'MA'           ,'skiprows':1,'header':0    ,'mag':412                    }                  ,\
     'OBS501'      :{'file':'camp','sheet':'Stations_Mean','skiprows':1,'header':0    ,'mag':'BS_OBS501_Global[FBU]'}                  ,\
     'SPM'         :{'file':'main','sheet':'SPM'          ,'skiprows':0,'header':[0,1],'mag':('SPM[g/m3]','Mean')   }                  ,\
     'HACH'        :{'file':'main','sheet':'turbidityHACH','skiprows':0,'header':[0,1],'mag':('globalMean[FNU]','Unnamed: 2_level_1') },\
     'LOVI'        :{'file':'main','sheet':'turbidityLOVI','skiprows':0,'header':[0,1],'mag':('globalMean[FNU]','Unnamed: 2_level_1') },\
     'phytoSpecies':{'file':'mainX','sheet':'phytoSpecies'},\
     'HS4'         :{'file':'mainX','sheet':'HS4'         },\
     'PSD'         :{'file':'mainX','sheet':'PSD'         },\
     'AC9'         :{'file':'mainX','sheet':'AC9'         },\
     'Mineralogy'  :{'file':'mainX','sheet':'Mineralogy'  },\
     'SUNPHOT'     :{'file':'specX','sheet':'SUNPHOT'     },\
     'CTD'         :{'file':'specX','sheet':'CTD'         }}


inventory = pd.DataFrame()
for camp in campaigns:
    region = camp.split('_')[0]
    date   = camp.split('_')[1]
    pathCampaign  = path0 + '/regions/' + region
    
    inventory.loc[camp,'region'] = region
    inventory.loc[camp,'date'  ] = date
    for m in M:
        try:
            if   M[m]['file'] == 'rad' :
                pathXls = glob.glob(pathCampaign + '/' + camp + '/' + m + 'Processed/' + region + '_' + date + '_' + m + '_QC_RhowStd*' + 'SatSensors.xlsx')[0]
            elif M[m]['file'] == 'camp':
                pathXls = pathCampaign + '/' + camp + '/campbellProcessed/' + region + '_' + date + '_Campbell.xlsx'
            elif M[m]['file'] == 'main' or M[m]['file'] == 'mainX':
                pathXls = pathCampaign + '/' + camp + '/' + region + '_' + date + '.xlsx'
            elif M[m]['file'] == 'specX':
                pathXls = pathCampaign + '/' + camp + '/' + region + '_' + date + '_' + m + '.xlsx'
            if M[m]['file'][-1] != 'X':
                ST = pd.read_excel(pathXls, sheet_name=M[m]['sheet'], skiprows=M[m]['skiprows'],header=M[m]['header'])[M[m]['mag']]
                numSt = np.sum(~ np.isnan(ST))
            else:
                ST = pd.read_excel(pathXls, sheet_name=M[m]['sheet'])
                numSt = 'X'
        except:
            numSt = ''
        if numSt == 0:
            numSt = ''
        inventory.loc[camp,m] = numSt
inventory.index.name = 'CampaignID'

wb = openpyxl.Workbook()
wb.save(pathXlsx)
writer = pd.ExcelWriter(pathXlsx, engine = 'openpyxl')
writer.book = wb
del wb['Sheet']

# Write to Excel: scalars0
sheetname = 'inventory'
if sheetname in wb.sheetnames:
    del wb[sheetname]
inventory.to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)

writer.save()
writer.close()

#%% Instruments per campaign

#pathRegions   = path0 + '/regions'
#pathXlsx      = path0 + '/general' + '/campaignsInventory.xlsx'
#
##Tagus_20180828 RdP_20121113_Muelle_SeaSWIR
#
#
#M = {'HACH'        :['turbidityHACH'     ,'mainSheet'],\
#     'SPM'         :['SPM'               ,'mainSheet'],\
#     'CHL'         :['CHL'               ,'mainSheet'],\
#     'phytoSpecies':['phytoSpecies'      ,'mainSheet'],\
#     'HS4'         :['HS4'               ,'mainSheet'],\
#     'PSD'         :['PSD'               ,'mainSheet'],\
#     'AC9'         :['AC9'               ,'mainSheet'],\
#     'Mineralogy'  :['Mineralogy'        ,'mainSheet'],\
#     'LOVI'        :['turbidityLOVIBAND' ,'mainSheet'],\
#     'ASD'         :['ASDProcessed'      ,'folders'  ],\
#     'TriOS'       :['TriosProcessed'    ,'folders'  ],\
#     'OBS'         :['campbellContinuous','folders'  ],\
#     'SUNPHOT'     :['SUNPHOT'           ,'xtraSheet'],\
#     'CTD'         :['CTD'               ,'xtraSheet']}
#
#
#inventory = pd.DataFrame()
#for camp in campaigns:
#    region = camp.split('_')[0]
#    date   = camp.split('_')[1]
#    pathCampaign  = path0 + '/regions/' + region
#    
#    inventory.loc[camp,'region'] = region
#    inventory.loc[camp,'date'  ] = date
#    for m in M:
#        cond = False
#        if   M[m][1] == 'mainSheet':
#            try:
#                pd.read_excel(pathCampaign + '/' + camp + '/' + region + '_' + date + '.xlsx', sheet_name=M[m][0], skiprows=1)
#                cond = True
#            except:
#                pass
#        elif M[m][1] == 'folders':
#            if  os.path.isdir(pathCampaign + '/' + camp + '/' + M[m][0]):
#                cond = True
#        elif M[m][1] == 'xtraSheet':
#            if os.path.isfile(pathCampaign + '/' + camp + '/' + region + '_' + date + '_' + M[m][0] + '.xlsx'):
#                cond = True
#        if cond:
#            inventory.loc[camp,m] = 'X'
#        else:
#            inventory.loc[camp,m] = ''
#
#inventory.index.name = 'CampaignID'
#
#wb = openpyxl.Workbook()
#wb.save(pathXlsx)
#writer = pd.ExcelWriter(pathXlsx, engine = 'openpyxl')
#writer.book = wb
#del wb['Sheet']
#
## Write to Excel: scalars0
#sheetname = 'inventory'
#if sheetname in wb.sheetnames:
#    del wb[sheetname]
#inventory.to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)
#
#writer.save()
#writer.close()
#%% Change input text files
for camp in campaigns:
    region = camp.split('_')[0]
    date   = camp.split('_')[1]
    pathCampaign  = path0 + '/regions/' + region
    filename = pathCampaign + '/' + camp + '/ASD/asdProcessingInputs'

############ METODO 1    
    text_to_search   = 'RhowStdThresh       0.005,0.005,0.005,0.005'
    replacement_text = 'RhowStdThresh       0.01,0.01,0.01,0.01'

    try:
        with fileinput.FileInput(filename, inplace=1) as file:
            for line in file:
                print(line.replace(text_to_search, replacement_text), end='')
    except:
        print(camp + ': No ASD-input file!')

############ METODO 2
#    fileRdP = path0 + '/regions/RdP/RdP_20141122_Samborombon/ASD/asdProcessingInputs'
#    try:
#        shutil.copyfile(fileRdP,filename)
#    except:
#        print(camp + ': No ASD-input file!')
#%% 
for camp in campaigns:
    region = camp.split('_')[0]
    date   = camp.split('_')[1]
    pathCampaign  = path0 + '/regi ons/' + region
    filename = pathCampaign + '/' + camp + '/ASD/asdProcessingInputs_conae1'
    if os.path.isfile(filename):
        print(camp)

#%% Reflectance-derived products
#def reflectanceDerived(campaign0,path0):
#    pass
#%%
#def scalarMeasurementsStation(campaign0,magnitudes,path0):
#
#    region = campaign0.split('_')[0]
#    month  = campaign0.split('_')[1]
#    pathCampaign  = pathRegions + '/' + region + '/' +  campaign0
#    campaign = region + '_' + month
#
#    pathRegions   = path0 + '/regions'
#    pathCampaign  = pathRegions + '/' + region + '/' +  campaign
#
#    stationInfo = pd.read_excel(pathCampaign + '/' + region + '_' + month + '.xlsx', sheetname='stationInfo', skiprows=1)
#    stationInfo['StationID'] = campaign + '_' + stationInfo['StationID'].astype(str)
#    stationInfo = stationInfo.set_index('StationID')
#
#    ## PLOTS PER STATIONS
#    M = {'HACH'     :{'xlsxFileName': pathCampaign + '/' + region + '_' + month + '.xlsx'                            ,\
#                      'sheetName'   : {'Mean': 'turbidityHACH'        , 'CV': 'turbidityHACH'        }               ,\
#                      'header'      : {'Mean': ('globalMean[FNU]','') , 'CV': ('globalCV[%]','')     }               ,\
#                      'skiprows'    : 0                                                                             },\
#
#        'SPM'       :{'xlsxFileName': pathCampaign + '/' + region + '_' + month + '.xlsx'                            ,\
#                      'sheetName'   : {'Mean': 'SPM'                  , 'CV': 'SPM'                  }               ,\
#                      'header'      : {'Mean': ('SPM[g/m3]','Mean')   , 'CV': ('SPM[g/m3]','CV[%]')  }               ,\
#                      'skiprows'    : 0                                                                             },\
#
#        'SOM'       :{'xlsxFileName': pathCampaign + '/' + region + '_' + month + '.xlsx'                            ,\
#                      'sheetName'   : {'Mean': 'SPM'                  , 'CV': 'SPM'                  }               ,\
#                      'header'      : {'Mean': ('SOM[g/m3]','Mean')   , 'CV': ('SOM[g/m3]','CV[%]')  }               ,\
#                      'skiprows'    : 0                                                                             },\
#
#        'BS_OBS501' :{'xlsxFileName': pathCampaign + '/campbellProcessed/' + region + '_' + month + '_Campbell.xlsx' ,\
#                      'sheetName'   : {'Mean': 'Stations_Mean'        , 'CV': 'Stations_CV'          }               ,\
#                      'header'      : {'Mean': 'BS_OBS501_Global[FBU]', 'CV': 'BS_OBS501_Global[FBU]'}               ,\
#                      'skiprows'    : 1                                                                             },\
#
#        'SS_OBS501' :{'xlsxFileName': pathCampaign + '/campbellProcessed/' + region + '_' + month + '_Campbell.xlsx' ,\
#                      'sheetName'   : {'Mean': 'Stations_Mean'        , 'CV': 'Stations_CV'          }               ,\
#                      'header'      : {'Mean': 'SS_OBS501_Global[FNU]', 'CV': 'SS_OBS501_Global[FNU]'}               ,\
#                      'skiprows'    : 1                                                                             },\
#       }
#
#    # Initialize magnitude vector
#    magStats   = {'Mean': pd.DataFrame(index=stationInfo.index),\
#                  'CV'  : pd.DataFrame(index=stationInfo.index)}
#    for stat in ['Mean','CV']:
#        for m in magnitudes:
#            assert(m in M.keys())
#            try:
#                if isinstance(M[m]['header'][stat], tuple):
#                    headerLines = list(range(0,len(M[m]['header'][stat])))
#                else:
#                    headerLines = [0]
#                headerRedefUnnamed = []
#                M[m][stat] = pd.read_excel(M[m]['xlsxFileName'],header=headerLines, sheetname=M[m]['sheetName'][stat],skiprows=M[m]['skiprows'])
#                for head in headerLines:
#                    headerRedef = M[m][stat].columns.get_level_values(level=head).str.replace('Un.*','')
#                    headerRedefUnnamed.append(headerRedef)
#                M[m][stat].columns = headerRedefUnnamed
#                mag0 = M[m][stat][M[m]['header'][stat]]
#                # Solving BUG in pandas (if first station of a 'multiheader' sheet is empty, it doesn't take it as part of the stations)
#                if len(mag0) != len(stationInfo):
#                    mag0 = pd.concat([pd.Series([np.nan],index=[mag0.index.name]), mag0])
#                # Change IDs and name to append to a general DF
#                mag0.index = stationInfo.index.tolist()
#                mag0.name  = m
#                magStats[stat] = pd.concat([magStats[stat],mag0        ],axis=1)
#            except:
#                if stat == 'Mean':
#                    print('No ' + m + ' measurements for this campaign!')
#                continue
#        magStats[stat] = pd.concat([stationInfo, magStats[stat]],axis=1)
#
#
#    return magStats['Mean'], magStats['CV']
#%%
#def scalarMeasurements(campaigns,magnitudes,path0):
#    campaigns = campaignList(campaigns,path0)
#    Mean = pd.DataFrame()
#    CV   = pd.DataFrame()
#    for campaign in campaigns:
#        print('Campaign: ' + campaign)
#        MeanCamp, CVCamp=scalarMeasurementsStation(campaign0,magnitudes,path0)
#        Mean = pd.concat([Mean,MeanCamp],axis=0)
#        CV   = pd.concat([CV  ,CVCamp  ],axis=0)
#    return Mean, CV
#%%
#def scalarMeasurementsPlot(Mean,CV,magnitudes):
#
##ax2 = Mean.plot.scatter(x='SPM',y='SOM',c='Place',colormap='viridis')
#
#groupsMean = {k:v for k,v in list(Mean.groupby(['Place','Region']))}
#groupsCV   = {k:v for k,v in list(  CV.groupby(['Place','Region']))}
#
#
#magPairs = list(itertools.combinations(magnitudes,2))
#
## Plot
#for pair in magPairs:
#    fig, ax = plt.subplots()
#    ax.margins(0.05) # Optional, just adds 5% padding to the autoscaling
#    for group in groupsMean.keys():
#        magX = pair[0]
#        magY = pair[1]
#        name = group[0] + ' (' + group[1] + ')'
#    #    ax.plot(groupsMean[group][magX], groupsMean[group][magY], marker='.', linestyle='', ms=12, label=name)
#        xerr = groupsMean[group][magX]*groupsCV[group][magX]/100
#        yerr = groupsMean[group][magY]*groupsCV[group][magY]/100
#        ax.errorbar(groupsMean[group][magX], groupsMean[group][magY], xerr=xerr, yerr=yerr, fmt='o', label=name)
#        plt.xlabel(magX)
#        plt.ylabel(magY)
#    ax.legend()
#    plt.show()
#


#%%
#def scalarMeasurementsStationProcess(campaign0,path0):
#    pathRegions   = path0 + '/regions'
#
#    region = campaign0.split('_')[0]
#    month  = campaign0.split('_')[1]    
#
#    campaign = region + '_' + month
#    pathCampaign  = pathRegions + '/' + region + '/' +  campaign
#    
#    magnitudes = ['HACH','SPM','SOM','BS_OBS501','SS_OBS501']
#
#    Mean, CV = scalarMeasurementsStation(campaign0,magnitudes,path0)
#
#    if not os.path.isdir(pathCampaign + '/scalarMeasurements/'):
#        os.mkdir(pathCampaign + '/scalarMeasurements/')
#    
#    figPath = pathCampaign + '/scalarMeasurements/'
#    if not os.path.isdir(figPath):
#        os.mkdir(figPath)
#        os.mkdir(figPath + 'png/')
#        os.mkdir(figPath + 'pdf/')
#
#    # Write to Excel: General TriOS workbook
#
#    pathXlsx = pathCampaign + '/TriosProcessed/' + campaign + '_Trios' + qc750Flag[1] + '.xlsx'
#    
#    wb = openpyxl.Workbook()
#    wb.save(pathXlsx)
#    
#    writer = pd.ExcelWriter(pathXlsx, engine = 'openpyxl')
#    writer.book = wb
#    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))    
#    
#    # Write to Excel:
#    sheetname = 'stationInfo'
#    if sheetname in wb.sheetnames:
#        wb.remove_sheet(wb.get_sheet_by_name(sheetname))
#    stationInfo.set_index('StationID')
#    stationInfo.to_excel(writer, index = False, sheet_name=sheetname,startrow = 1)
#%% Intercompare turbidity/MPS measurements

#campaigns = 'all'
#path0  = '/home/gossn/Dropbox/Documents/Data/inSitu/juanchoProcess'
#magnitudes = ['HACH','SPM','SOM','BS_OBS501','SS_OBS501']
#
#
#Mean, CV = scalarMeasurements(campaigns,magnitudes,path0)

#%% Extract TriOS measurements
#def triosMeasurements(campaigns,magnitudes,path0):
#    campaigns = campaignList(campaigns,path0)
#    Mean = pd.DataFrame()
##    CV   = pd.DataFrame()
#    for campaign in campaigns:
#        print('Campaign: ' + campaign)
#        MeanCamp, CVCamp=scalarMeasurementsStation(campaign0,magnitudes,path0)
#        Mean = pd.concat([Mean,MeanCamp],axis=0)
#        CV   = pd.concat([CV  ,CVCamp  ],axis=0)
#    return Mean
#%% CAMPAÑA SEASWIR: RdP_20121113_Muelle_SeaSWIR: Obtener datos convolucionados

#OJO!!! SOBREESCRIBE LAS HOJAS PREEXISTENTES

import pandas as pd
# Read SRFs
sensors = ['MA','MT','ME','MSI','OLI','OLCI','PHR1B','VIIRS']
srfs = {}
for sensor in sensors:
        srfs[sensor] = pd.read_excel('/home/gossn/Dropbox/Documents/inSitu/Database/SRFs/SpectralResponse_' + sensor + '.xlsx').set_index('lambdas')
        srfs[sensor] = srfs[sensor].fillna(0)
        srfs[sensor][srfs[sensor]<0] = 0

pathXlsx = '/home/gossn/Dropbox/Documents/inSitu/Database/regions/RdP/RdP_20121113_Muelle_SeaSWIR/ASDProcessed/prueba'

wb = openpyxl.Workbook()
writer = pd.ExcelWriter(pathXlsx + '2.xlsx', engine = 'openpyxl')
writer.book = wb


signal = pd.read_excel(pathXlsx + '.xlsx',sheet_name='Rhow',skiprows=1)
signal = signal.set_index('Unnamed: 0')
signal.index.names = ['']
wavelengths = np.array(signal.index)
signal = signal.T

dsignal = pd.read_excel(pathXlsx + '.xlsx',sheet_name='RhowStd',skiprows=1)
dsignal = dsignal.set_index('Unnamed: 0')
dsignal.index.names = ['']
dsignal = dsignal.T

RhowBands,dRhowBands,lambdaSrfsMean=cp0.convoluteSrfs(wavelengths,signal,dsignal,srfs,sensors,writer,wb)

lambdaSrfs      = {}
lambdaSrfsMean  = {}
RhowLambdaSrfs  = {}
dRhowLambdaSrfs = {}
RhowBands       = {}
dRhowBands      = {}

for sensor in sensors:
    lambdaSrfs     [sensor] = srfs[sensor].index
    lambdaSrfsMean [sensor] = {}
    RhowLambdaSrfs [sensor] = pd.DataFrame(columns=lambdaSrfs[sensor])
    dRhowLambdaSrfs[sensor] = pd.DataFrame(columns=lambdaSrfs[sensor])
    RhowBands      [sensor] = pd.DataFrame(columns=srfs[sensor].columns)
    dRhowBands     [sensor] = pd.DataFrame(columns=srfs[sensor].columns)

    # calculate mean wavelength
    for band in srfs[sensor].columns:
        lambdaSrfsMean[sensor][band] = np.sum(list(lambdaSrfs[sensor])*srfs[sensor][band])/np.sum(srfs[sensor][band])

    # interp Rhows to lambdaSrfs:
    for st in signal.index:
#            if np.all(signal.loc[st].isna()):
#                RhowLambdaSrfs[ sensor].loc[st,:] = np.nan
#                dRhowLambdaSrfs[sensor].loc[st,:] = np.nan
#                RhowLambdaSrfs[ sensor] df.append(pandas.Series(), ignore_index=True)
        RhowLambdaSrfs[ sensor].loc[st,:] = np.interp(lambdaSrfs[sensor], wavelengths,  signal.loc[st])
        dRhowLambdaSrfs[sensor].loc[st,:] = np.interp(lambdaSrfs[sensor], wavelengths, dsignal.loc[st])
        for band in srfs[sensor].columns:
            if wavelengths[0] <= lambdaSrfsMean[sensor][band] <= wavelengths[-1]:
                S = np.sum(srfs[sensor][band])
                RhowBands[sensor].loc[ st,band] =         np.sum(  RhowLambdaSrfs[sensor].loc[st,:]*srfs[sensor][band])/S
                dRhowBands[sensor].loc[st,band] = np.sqrt(np.sum((dRhowLambdaSrfs[sensor].loc[st,:]*srfs[sensor][band]/S)**2))
            else:
                RhowBands[ sensor].loc[st,band] = np.nan
                dRhowBands[sensor].loc[st,band] = np.nan

    if excelWriter != 'noExcel':
        sheetname = sensor
        if sheetname in wb.sheetnames:
            wb.remove_sheet(wb.get_sheet_by_name(sheetname))
        RhowBands[sensor].to_excel(excelWriter, index = True, sheet_name=sheetname,startrow = 1)
#            adjustColWidth(wb.get_sheet_by_name(sheetname))
        excelWriter.save(); excelWriter.save()
        excelWriter.close(); excelWriter.close()

        sheetname = sensor + 'Std'
        if sheetname in wb.sheetnames:
            wb.remove_sheet(wb.get_sheet_by_name(sheetname))
        dRhowBands[sensor].to_excel(excelWriter, index = True, sheet_name=sheetname,startrow = 1)
#            adjustColWidth(wb.get_sheet_by_name(sheetname))
        excelWriter.save(); excelWriter.save()
        excelWriter.close(); excelWriter.close()