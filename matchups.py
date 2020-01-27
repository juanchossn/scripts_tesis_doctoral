#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Sep 12 10:48:17 2019

@author: gossn
"""
#%%
import sys
import os
import pandas as pd
import numpy as np
from netCDF4 import Dataset
import matplotlib.pyplot as plt
import ast
import glob
import math
import openpyxl 
from openpyxl import load_workbook
import ast


# This home-made module is just used to plot the RGBs, but I'm sure you will 
# find a better way of plotting the RGBS. So I would simply comment these (3)
# lines...
pathHomeMadeModules = '/home/gossn/Dropbox/Documents/pyHomeMadeModules'
pathDB       = '/home/gossn/Dropbox/Documents/inSitu/Database/matchUpImages'
pathMatchUps = '/home/gossn/Dropbox/Documents/inSitu/Database/general/'

sys.path.append(pathHomeMadeModules)
import roiNcdf as rn
#%% Function: flag_data_fast
def flag_data_fast(sensor):
    '''This function returns a function (a priori, sensor-dependent) that 
    computes a mask resulting from the juxtaposition of the pertinent flags'''
    # You can change the flags / path to flags as you require
    def main(pathImg):
        if   sensor == 'MA' or sensor == 'MT':# MA/MT: MODIS Aqua/Terra
            pathFlags = pathImg + '/' + pathImg.split('/')[-1] + '.L2_RC'
            flags_we_want = ['LAND','CLDICE', 'HISATZEN', 'HISOLZEN']
            flags = Dataset(pathFlags,'r')['geophysical_data']['l2_flags']
            flag_bits = np.uint32()
        elif sensor == 'OLCI':
            pathFlags = pathImg + '/' + 'rhos'
            flags_we_want = ['LAND','CLDICE', 'HISATZEN', 'HISOLZEN']
            flags = Dataset(pathFlags,'r')['geophysical_data']['l2_flags']
            flag_bits = np.uint32()
#            pathFlags = pathImg + '/' + pathImg.split('/')[-1] + '.L2_RC'
#            flags_we_want = ['CLOUD','LAND', 'CLOUD_AMBIGUOUS', 'CLOUD_MARGIN',   \
#                             'INVALID', 'COSMETIC', 'SATURATED', 'SUSPECT',       \
#                             'HISOLZEN', 'HIGHGLINT', 'SNOW_ICE', 'AC_FAIL',      \
#                             'WHITECAPS']
#            flags = Dataset(pathFlags + '/wqsf.nc','r')['WQSF']
#            flag_bits = np.uint64()
        elif sensor == 'MSI':
            pass # you can perform the function for MSI here ;)
        flag_data     = flags[:,:]
        flag_names    = flags.flag_meanings.split(' ')
        flag_values   = flags.flag_masks
        
        for flag in flags_we_want:
            try:
                flag_bits = flag_bits | flag_values[flag_names.index(flag)]
            except:
                print(flag + " not present")
        flag_mask = (flag_data & flag_bits) > 0
        flag_mask = flag_mask.astype(float)
        flag_mask[flag_mask == 0.0] = np.nan
        flag_mask = np.isfinite(flag_mask)
        return flag_mask
    return main
#%%
def latlon(sensor):
    '''This function returns a function (a priori, sensor-dependent) that 
    computes the lat lon matrices required to search the match-up window'''
    # You can add the specific function for MSI 
    # You will probably need to edit this function according to the way you store your processed images
    def main(pathImg):
        if   sensor == 'MA' or sensor == 'MT':
            pathLatLon = pathImg + '/' + pathImg.split('/')[-1] + '.L2_RC'
            lat = Dataset(pathLatLon)['navigation_data']['latitude' ][:,:]
            lon = Dataset(pathLatLon)['navigation_data']['longitude'][:,:]
        elif sensor == 'OLCI':
            pathLatLon = pathImg + '/geoBlr'
            lat = Dataset(pathLatLon)['lat'][:,:]
            lon = Dataset(pathLatLon)['lon'][:,:]
        elif sensor == 'MSI':
            pass
        return lat,lon
    return main

#%% 
def rasterSensors(sensor,rasterID):
    '''This function will return a sensor- and raster-dependent function that 
    will return a dictionary with the bands of the rasters from which you 
    desire to extract '''
    # You can add the specific function for MSI 
    # You will probably need to edit this function according to the way you store your processed images
    if sensor == 'MA' or sensor == 'MT':
        rDir = 'L2'
    if sensor == 'OLCI':
        rDir = 'DER'
    def main(pathImg,flags):
        raster = {}
        if   sensor == 'MA' or sensor == 'MT':
            pathRaster = pathImg + '/' + pathImg.split('/')[-1] + '.L2_' + rasterID
            pathRC     = pathImg + '/' + pathImg.split('/')[-1] + '.L2_RC'
            bands = [412, 443, 469, 488 , 531 ,\
                     547, 555, 645, 667 , 678 ,\
                     748, 859, 869, 1240, 1640, 2130]
            if rasterID == 'RC':
                prod0 = 'rhos_'
            elif 'PCA' in rasterID:
                prod0 = 'rhow'
                bands = [b for b in bands if b not in [1240, 1640, 2130]]
            else:
                prod0 = 'Rrs_'
            prod = [prod0 + str(b) for b in bands]
            for p in prod:
                if 'PCA' in rasterID:
                    raster[p] = Dataset(pathRaster)[p][:,:]
                else:
                    raster[p] = Dataset(pathRaster)['geophysical_data'][p][:,:]
                raster[p][flags] = np.nan
        elif sensor == 'OLCI':
            pathRC     = pathImg + '/rhos'
            if rasterID == 'BLR':
                pathRaster = pathImg + '/rhoWBlr'
                prod0 = 'rhow'
                bands = [620,665,709,779,865,1016]
                prod = [prod0 + '_' + str(b) for b in bands]
                for p in prod:
                    raster[p] = Dataset(pathRaster)[p][:,:]
                    raster[p][flags] = np.nan
            if rasterID == 'RC':
                pathRaster = pathImg + '/rhos'
                prod0 = 'rhos'
                bands = [400,412,442,490,510,560,620,665,\
                         674,681,709,754,761,764,768,779,\
                         865,885,900,940,1012]
                prod = [prod0 + '_' + str(b) for b in bands]
                for p in prod:
                    raster[p] = Dataset(pathRaster)['geophysical_data'][p][:,:]
                    raster[p][flags] = np.nan
        elif sensor == 'MSI':
            pass
        return raster,pathRC
    return main,rDir
#%% 
def matchUpStations(pathMatchUps,region,sensor):
    '''This function will return a dataframe with all the pertinent information
    of each in situ station to perform match-up. I will send you the Excel that
    I've created that is used as input here. Take care of the path!'''
    matchUps = pd.read_excel(pathMatchUps + 'overpasses.xlsx',sheet_name='stationInfo',skiprows=1)
    matchUps = matchUps.set_index('StationID')
    overpasses = matchUps['Overpasses']
    for st in overpasses.index:
        string = overpasses.loc[st]
        acceptableString = string.replace('‘', '\'').replace('’', '\'')
        overpasses = overpasses.replace(string, acceptableString)
    matchUps['Overpasses'] = overpasses
    if region == 'all':
        stationsRegion = matchUps.Region != ''
    else:
        stationsRegion = matchUps.Region == region
    if not np.any(stationsRegion):
        ValueError('Non-existent region!')

    stationsSensor = pd.Series()
    for st in matchUps.index:
        if sensor in ast.literal_eval(matchUps.loc[st,'Overpasses']):
            stationsSensor.loc[st] = True
        else:
            stationsSensor.loc[st] = False

    matchUps = matchUps.loc[stationsRegion & stationsSensor,:]
    return matchUps

#%% In situ DB
def extractWindows(region,sensor,rasterID,pathDB,pathMatchUps,RGBFlag,minimumFlag):
    '''This function will create a dataframe that will be later stored as an 
    Excel file. Each Excel will have many sheets, three per window size 
    (each of the three being Mean, Std and # of Valid pixels inside window'''

#    region       = 'Tagus'
#    pathDB       = '/home/gossn/Dropbox/Documents/inSitu/Database/matchUpImages'
#    pathMatchUps = '/home/gossn/Dropbox/Documents/inSitu/Database/general/'
#    sensor       = 'MA'
#    rasterID     = 'RC'
#    RGBFlag      = False
#    minimumFlag  = True


    Minimums = pd.DataFrame(columns=['Row','Col'])
    
    matchUps = matchUpStations(pathMatchUps,region,sensor)
    
    if len(matchUps.index) == 0:
        print('No ' + sensor + ' matchups for ' + region + '!')
        return
    latlonSensor = latlon(sensor)
    masks        = flag_data_fast(sensor)
    rasters,rDir = rasterSensors(sensor,rasterID)
    
    winSizeParam = np.arange(0,4)
    rd = math.pi/180
    
    dfFlag = True
    match  = {}
    #   PRODUCT
    st0 = -1; S   = len(matchUps)

    
    for st in matchUps.index:
        st0+=1
        print(sensor + '; ' + region + '; ' + rasterID + ': ' + str(np.round(st0/S*100)) + ' %')
        img0 = ast.literal_eval(matchUps.loc[st,'Overpasses'])
        img = img0[sensor]
        
        path0 = pathDB + '/' + sensor + '/' + region + '/matchUps'
        
        pathImg       = path0 + '/' + rDir + '/' + img
        lat,lon       = latlonSensor(pathImg)
        maskImg       = masks(pathImg)
        raster,pathRC = rasters(pathImg,maskImg)
    
    
        stIS = matchUps.loc[st,['Lat','Lon']]
        
        if np.isnan(stIS['Lat']):
            continue
        # Compute orthodromic angle (we did this in Lisbon)
        dS = np.arccos(np.sin(lat*rd)*np.sin(stIS['Lat']*rd) + \
                       np.cos(lat*rd)*np.cos(stIS['Lat']*rd)*np.cos((lon-stIS['Lon'])*rd))

        # Take the closest pixel to station's location
        M = np.unravel_index(dS.argmin(), np.shape(dS))
        
        Minimums.loc[st,:] = [M[0],M[1]]
        
        '''
        !!!! This 'for' block will create RGBs with the images where you have matchups
        It won't work for you, although I will send you the outputs,
        if you wish to make it work, speak to me.
        '''

        if RGBFlag: # You should switch this to False (at the final block inside the script)
            projectFlag = False
            plotFlag    = False
            if sensor == 'MA' or sensor == 'MT':
                RCsoftware  = 'seadas'
                satRGB   = 0.12
                degStep  = 0.005
                RGBBands = [645, 555, 488]
            if sensor == 'OLCI':
                RCsoftware  = 'seadas'
                satRGB   = 0.12
                degStep  = 0.005
                RGBBands = [620,560,442]
            try:
                os.mkdir(path0 + '/RGB')
            except:
                pass


            figOut   = path0 + '/RGB/' + st + '_' + img + '_RGB'

            for z in ['NO_ZOOM','ZOOM']:
                rgb = rn.rc2rgb(pathRC,figOut,'noROI',projectFlag,plotFlag,RGBBands,RCsoftware,sensor,satRGB,degStep)
                for col in range(3):
                    rgb['val'][:,:,col] = np.multiply(rgb['val'][:,:,col],~maskImg)
                fig = plt.figure(figsize=(30.0, 25.0))
                plt.imshow(rgb['val'   ],origin='lower')#,extent=[boxRoi['W'],boxRoi['E'],boxRoi['S'],boxRoi['N']])
                plt.plot(M[1],M[0],'xg',mew=30)
                plt.xlabel(rgb['xlabel'],fontsize=18)
                plt.ylabel(rgb['ylabel'],fontsize=18)
                plt.title('IMG = ' + img + '; ST = ' + st,fontsize=18)
                plt.rc('xtick',labelsize=18)
                plt.rc('ytick',labelsize=18)
                plt.locator_params(nbins=5)
                if z == 'ZOOM':
                    plt.xlim(M[1]-15,M[1]+15)
                    plt.ylim(M[0]-15,M[0]+15)
                    figOut = figOut + '_' + z
                fig.savefig(figOut + '.png')
                plt.close('all')

        for d in winSizeParam:
            for stats in ['Valid','Mean','Std']:
                sheet = 'Win' + str(d) + '_' + stats
                if dfFlag:
                    match[sheet] = pd.DataFrame(columns=list(raster.keys()))
                    
                for v in raster.keys():
                    win = raster[v][(M[0]-d):(M[0]+d+1),(M[1]-d):(M[1]+d+1)]
                    win = win.ravel()
                    win = win[win>-10]
                    if   stats == 'Valid':
                        val = np.sum(~np.isnan(win))
                        if val == 0 or np.ma.is_masked(val):
                            val = np.nan
                    elif stats == 'Mean' :
                        try:
                            val = np.nanmean(win)
                        except:
                            val = np.nan
                    elif stats == 'Std'  :
                        try:
                            val = np.nanstd(win)
                        except:
                            val = np.nan
                    if len(win) == 0:
                        val = np.nan

                    match[sheet].loc[st,v] = val
        dfFlag = False
    
    # This block will store the created dataframes into an Excel file:
    pathXlsx  = path0 + '/' + sensor + '_' + region + '_' + rasterID + '.xlsx'
    wb = openpyxl.Workbook()
    wb.save(pathXlsx)
    writer = pd.ExcelWriter(pathXlsx, engine = 'openpyxl')
    writer.book = wb
    del wb['Sheet']
    
    for sheetname in match.keys():
        if sheetname in wb.sheetnames:
            del wb[sheetname]
        match[sheetname].to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)
        writer.save()
        writer.close()

    if minimumFlag:
        # This block will store the created dataframes into an Excel file:
        pathXlsx  = path0 + '/' + sensor + '_' + region + '_Minimum_Location.xlsx'
        wb = openpyxl.Workbook()
        wb.save(pathXlsx)
        writer = pd.ExcelWriter(pathXlsx, engine = 'openpyxl')
        writer.book = wb
        del wb['Sheet']

        Minimums.to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)
        writer.save()
        writer.close()
#%% MAIN
def main(regions,sensorRaster,pathDB,pathMatchUps):
    for region in regions:
        for sensor in sensorRaster.keys():
            RGBFlag     = False # You should switch this to off
            minimumFlag = True
            rasterIDs = sensorRaster[sensor]

            for rasterID in rasterIDs:
                extractWindows(region,sensor,rasterID,pathDB,pathMatchUps,RGBFlag,minimumFlag)
                RGBFlag     = False
                minimumFlag = False
#%% This block will execute the match-ups using all the pre-defined functions!
#regions = ['RdP','Tagus']
#sensorRaster = {'OLCI':['RC','BLR']                          ,\
#                'MA'  :['NIR_GW94','NIR_ITER','NIR_SWIR','RC',\
#                        'PCA_1241_1628_2114','PCA_1628_2114' ,\
#                        'PCA_1241_2114','PCA_1241_1628']     ,\
#                'MT'  :['NIR_GW94','NIR_ITER','NIR_SWIR','RC',\
#                        'PCA_1241_1628_2114','PCA_1628_2114' ,\
#                        'PCA_1241_2114','PCA_1241_1628']      }
##sensorRaster = {'OLCI':['RC'],\
##                'MA'  :['RC'],\
##                'MT'  :['RC']}


#regions = ['RdP']
#sensorRaster = {'MA'  :['PCA_1241_1628_2114','PCA_1628_2114' ,\
#                        'PCA_1241_2114','PCA_1241_1628']     ,\
#                'MT'  :['PCA_1241_1628_2114','PCA_1628_2114' ,\
#                        'PCA_1241_2114','PCA_1241_1628']      }
#
#main(regions,sensorRaster,pathDB,pathMatchUps)