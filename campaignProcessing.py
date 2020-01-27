#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Dec 19 10:11:29 2018

@author: gossn
"""

#%% Modules
import pandas as pd
import numpy as np
pd.options.mode.chained_assignment = None
import openpyxl
from openpyxl import load_workbook
import os
import datetime
import pytz
from pysolar.solar import *
import time
from datetime import date, datetime, time, timedelta
from datetime import datetime
import glob
from xlrd import open_workbook
from shutil import copyfile
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf
#plt.ioff()
#matplotlib.use('Agg')
import itertools
import fnmatch
import ast
import shutil
#%% campaignList
def campaignList(campaign,path0):
    campaign0 = []
    if campaign == ['all']:
        pathRegions = path0 + '/regions'
        for reg in os.listdir(pathRegions):
            for camp in os.listdir(pathRegions + '/' + reg):
                campaign0.append(camp)
        campaign = campaign0
    elif len(campaign[0].split('_')) == 1: # "campaign" is a list of regions
        for region in campaign:
            pathRegion = path0 + '/regions/' + region
            for camp in os.listdir(pathRegion):
                campaign0.append(camp)
        campaign = campaign0
    else: # "campaign" is a list of campaigns
        for camp0 in campaign:
            region = camp0.split('_')[0]
            month  = camp0.split('_')[1]
            pathRegion = path0 + '/regions/' + region
            camp = [f for f in os.listdir(pathRegion) if camp0 in f]
            campaign0.append(camp[0])
    campaign  = sorted(campaign0)
    return campaign
#%% Pandas2Excel
def xldate_to_datetime(xldate):
    import datetime
    temp = datetime.datetime(1899, 12, 30)
    delta = datetime.timedelta(days=xldate)
    return temp+delta

def adjustColWidth(workbookSheet):
    for col in workbookSheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        workbookSheet.column_dimensions[column].width = adjusted_width
#%% pic2stationPic
def pic2stationPic(campaign0,path0):

#    campaign0 = 'Tagus_20190617'
#    path0     = '/home/gossn/Dropbox/Documents/inSitu/Database'
    
    pathRegions   = path0 + '/regions'
    
    region = campaign0.split('_')[0]
    month  = campaign0.split('_')[1]
    pathCampaign  = pathRegions + '/' + region + '/' +  campaign0
    campaign = region + '_' + month
    
    print(pathCampaign + '/Pics/PicsProcessingInputs')
    
    inputs = {}
    try:
        with open(pathCampaign + '/Pics/PicsProcessingInputs') as f0:
            for line in f0:
               (key, val) = line.split()
               inputs[key] = val.split(',')
               if len(inputs[key])==1:
                   inputs[key] = inputs[key][0]
    except:
        print('No Pics Input file or no Pics for this campaign!')
#        return
    

    dfCampaign = pd.read_excel(pathCampaign + '/' + campaign + '.xlsx',sheet_name='stationInfo',skiprows=1)
    stations = dfCampaign['StationID']
    # Remendar bug openpyxl (borra formatos preestablecidos)
    dfCampaign['startTimeUTC'] = pd.to_datetime(dfCampaign['startTimeUTC'],format= '%H:%M:%S' )
    dfCampaign['timeStampUTC'] = dfCampaign['DateUTC'] + pd.to_timedelta(dfCampaign['startTimeUTC']) + timedelta(hours=70*24*365.25-12)
    #
    pathPics = pathCampaign + '/' + 'Pics'
    pathPicsStations = pathCampaign + '/' + 'PicsStations'
    try:
        os.mkdir(pathPicsStations)
    except:
        print(pathPicsStations + ' already exists!')
    
    picList=os.listdir(pathPics)
    
    for st in stations:
        for pic in picList:
            picModTime = datetime.fromtimestamp(os.path.getmtime(pathPics + '/' + pic)-float(inputs['deltaUTC'])*60*60)#.strftime('%Y-%m-%d %H:%M:%S')
            deltaPics = abs(dfCampaign.loc[dfCampaign['StationID'] == st,'timeStampUTC'] - picModTime)
            timeDeltaCond = (deltaPics < pd.Timedelta(float(inputs['timeDeltaStationMin']), unit='m')).asobject[0]
            if timeDeltaCond:
                copyfile(pathPics + '/' + pic, pathPicsStations + '/' + st + '_' + pic[0:-len('.jpg')] + '.jpg')

#%% OBS Processing
def campbellContinuous2Stations(campaign0,path0):
#    path0 = '/home/gossn/Dropbox/Documents/inSitu/Database'
#    campaign0 = 'BALakes_20161215_Junin'

    pathRegions   = path0 + '/regions'

    region = campaign0.split('_')[0]
    date   = campaign0.split('_')[1]
    campaign = region + '_' + date
    pathCampaign  = pathRegions + '/' + region + '/' +  campaign0
    
    print('Processing Campbell data for campaign: ' + campaign0)
    
    inputs = {}
    try:
        with open(pathCampaign + '/campbellContinuous/campbellContinuousProcessingInputs') as f0:
            for line in f0:
               (key, val) = line.split()
               inputs[key] = val.split(',')
               if len(inputs[key])==1:
                   inputs[key] = inputs[key][0]
    except:
        print('No Campbell Input file or no Campbell measurements for this campaign!')
        return
        
    if not os.path.isdir(pathCampaign + '/campbellProcessed/'):
        os.mkdir(pathCampaign + '/campbellProcessed/')
    # OBS Processing
    
    if not os.path.isdir(pathCampaign + '/campbellContinuous'):
        print("No OBS data available for this campaign!")
    else:
        # Append sheet to Excel logsheet with Campbell Sci. continuous measurements...
        filenameCs = os.listdir(pathCampaign + '/campbellContinuous')
        filenameCs.remove('campbellContinuousProcessingInputs')
        
        # Read station IDs and Times
        stationInfo = pd.DataFrame()
        stationInfo = pd.read_excel(pathCampaign + '/' + campaign + '.xlsx',sheet_name='stationInfo',skiprows=1)
    
        # Remendar bug openpyxl (borra formatos preestablecidos)
        stationInfo['startTimeUTC'] = pd.to_datetime(stationInfo['startTimeUTC'],format= '%H:%M:%S' )
        stationInfo['timeStampUTC'] = stationInfo['DateUTC'] + pd.to_timedelta(stationInfo['startTimeUTC']) + timedelta(hours=70*24*365.25-12)
        
        stationIDs   = stationInfo['StationID'].asobject
        stationTimes = stationInfo['timeStampUTC'].asobject
    
    
        pathXlsx = pathCampaign + '/campbellProcessed/' + campaign + '_Campbell.xlsx'
        wb = openpyxl.Workbook()
        wb.save(pathXlsx)
        writer = pd.ExcelWriter(pathXlsx, engine = 'openpyxl')
        writer.book = wb
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))   
    
        # Write to Excel: StationInfo
        sheetname = 'stationInfo'
        if sheetname in wb.sheetnames:
            wb.remove_sheet(wb.get_sheet_by_name(sheetname))
        stationInfo.set_index('StationID')
        stationInfo.to_excel(writer, index = False, sheet_name=sheetname,startrow = 1)

        # Initialize dictionary of dataFrame that will store data per station (all dataloggers)
        csStations = {}
        for stat in ['Mean','Std','CV']:
            csStations[stat] = pd.DataFrame(index=stationIDs)
    
        for file in filenameCs:
            
            csCont = pd.DataFrame()
            csCont = pd.read_csv(pathCampaign + '/campbellContinuous' + '/' + file, header = [1,2])
    
            # Change column names
            colNamesOld = list(csCont.columns)
            colNamesNew = []
            for cname in colNamesOld:
                if cname[1][:len('Unnamed')] == 'Unnamed':
                    colNamesNew.append(cname[0])
                else:
                    colNamesNew.append(cname[0] + '[' + cname[1].replace(' ','') + ']')
            csCont.columns = colNamesNew
    
            # campbell Times
            csCont = csCont.drop(csCont.index[[0]])    # Drop extra-headers
            csCont['TIMESTAMP[TS]'] = pd.to_datetime(csCont['TIMESTAMP[TS]'])-timedelta(hours=float(inputs['deltaUTC']))
            csContTime = csCont['TIMESTAMP[TS]']
    
            
            ##### DATOS POR ESTACION
            print('Data per station')
            
            # campbell collected data
            csContData = pd.DataFrame()
            others = [c for c in csCont.columns if (c.lower()[:2] == 'wd' or c.lower()[:12] == 'stationnames' or c.lower()[:7] == 'temp_cr')]
            csContData = csCont.drop(['TIMESTAMP[TS]','RECORD[RN]','BattV[Volts]'] + others, axis=1)
            csContData = csContData.apply(pd.to_numeric, errors='coerce')
            csMeasures = list(csContData.columns.values)
        
            csStStats  = {}
            for st in stationIDs:
                print('Processing station ' + str(st))
                timeSt = stationTimes[stationIDs==st][0]
                timeDeltaSt = abs(csContTime-timeSt)<pd.Timedelta(float(inputs['timeDeltaStationMin']), unit='m')
                if any(timeDeltaSt):
                    csContSt = csContData.loc[timeDeltaSt,:]
                    Q1 = csContSt.quantile(0.25)
                    Q3 = csContSt.quantile(0.75)
                    IQR = Q3 - Q1
                    outliers = (csContSt < (Q1 - 1.5 * IQR)) |(csContSt > (Q3 + 1.5 * IQR))
                    csContSt[outliers] = np.nan
                    csStMed = csContSt.mean()
                    csStStd = csContSt.std()
                    csStCV  = csStStd/csStMed*100
                    
                    csStStats[st] = {'Mean': csStMed, 'Std': csStStd,'CV': csStCV}
                    
            for stat in ['Mean','Std','CV']:
                csStationsFile = pd.DataFrame(columns=csMeasures)
                for st in stationIDs:
                    try:
                        csStationsFile.loc[st]   = csStStats[st][stat]
                    except:
                        csStationsFile.loc[st]   = pd.Series(index=csMeasures)
                csStationsFile.fillna('')
                csStationsFile.index.name = 'StationID'
                csStations[stat] = pd.concat([csStations[stat], csStationsFile], axis=1)
                                

            ##### DATOS CRUDOS
            sheetname = file[:-4]
            if sheetname in wb.sheetnames:
                wb.remove_sheet(wb.get_sheet_by_name(sheetname))
            csCont.to_excel(writer, index = False, sheet_name=sheetname)
            #adjustColWidth(wb.get_sheet_by_name(sheetname))
            writer.save()
            writer.close()
            
            ##### DATOS SUAVIZADOS A 'smoothWinMin' MIN
            if file[0:6] != 'CR200X':
                csSmooth = pd.DataFrame()
                step = 0
                flag = True
                while flag:
                    timeWin = csCont['TIMESTAMP[TS]'].min() + timedelta(minutes=step*float(inputs['smoothWinMin']))
                    step+=1
                    flag = timeWin < csCont['TIMESTAMP[TS]'].max()
                    timeDeltaWin = abs(csContTime-timeWin)<pd.Timedelta(float(inputs['smoothWinMin'])/2, unit='m')
                    if any(timeDeltaWin):
                        csContWin = csContData.loc[timeDeltaWin,:]
                        Q1 = csContWin.quantile(0.25)
                        Q3 = csContWin.quantile(0.75)
                        IQR = Q3 - Q1
                        outliers = (csContWin < (Q1 - 1.5 * IQR)) |(csContWin > (Q3 + 1.5 * IQR))
                        csContWin[outliers] = np.nan
                        csWinMedia = csContWin.mean()
                        csWinStd = csContWin.std()
                        csWinCV = csWinStd/csWinMedia*100
                        
#                        csWinStats = [csWinMedia,csWinStd,csWinMedia]
                        
                        csMeasures = list(csContWin.columns.values)
                        for m in csMeasures:
                            csSmooth.loc[timeWin,m + 'Mean'  ] = csWinMedia[m]
                            csSmooth.loc[timeWin,m + 'Std'  ] = csWinStd[m]
                            csSmooth.loc[timeWin,m + 'CV'] = csWinCV[m]
                sheetname = file[:-4] + 'ContSmooth' + inputs['smoothWinMin'] + 'min'
                if sheetname in wb.sheetnames:
                    wb.remove_sheet(wb.get_sheet_by_name(sheetname))
                csSmooth.to_excel(writer, index = True, sheet_name=sheetname)
                #adjustColWidth(wb.get_sheet_by_name(sheetname))
                writer.save()
                writer.close()


        ##### GLOBAL VARIABLES: compute
        globalVars = ['BS_OBS501_','SS_OBS501_']
        globVar = {}
        for var in globalVars:
            colsDf  = [c for c in csStations[stat].columns if c.startswith(var)]
            unitsDf = '[' + colsDf[0].split('[')[1]
            # Compute MU
            globVar['Mean'] = csStations['Mean'][colsDf].mean(axis=1)
            # Compute composite STD asumming equal weights: n1=n2=...
            # sigma**2 = 0.5*(sigma1**2 + sigma2**2 + ... + mu1**2 + mu2**2 + ... -N*muGlobal**2)
            N = csStations['Mean'][colsDf].notnull().sum(axis=1)
            globVar['Std'] = pd.Series(index=stationIDs)
            for col in colsDf:
                globVar['Std'] = globVar['Std'].add(csStations['Mean'][col]**2,fill_value=0)
                globVar['Std'] = globVar['Std'].add(csStations['Std'][col]**2,fill_value=0)
            globVar['Std'] = globVar['Std'].add(-N*(globVar['Mean']**2),fill_value=0)
            globVar['Std'] = (globVar['Std']/N)**0.5
            # Compute CV
            globVar['CV'] = 100*globVar['Std']/globVar['Mean']
            # append computed global variables to csStations
            for stat in ['Mean','Std','CV']:
                csStations[stat][var + 'Global' + unitsDf] = globVar[stat]

            var = 'BS_OBS501'  
            csStations['Mean'][[c for c in csStations[stat].columns if c.startswith(var)]]

        for stat in ['Mean','Std','CV']:
            csStations[stat] = csStations[stat].reindex(columns=sorted(csStations[stat].columns))
            sheetname = 'Stations' + '_' + stat
            if sheetname in wb.sheetnames:
                wb.remove_sheet(wb.get_sheet_by_name(sheetname))
            csStations[stat].to_excel(writer, index = True, sheet_name=sheetname, startrow=1)
            #adjustColWidth(wb.get_sheet_by_name(sheetname))
            writer.save()
            writer.close()
#%% convoluteSrfs
def convoluteSrfs(wavelengths,signal,dsignal,srfs,sensors,excelWriter,wb):

    #RhowBands,dRhowBands,lambdaSrfsMean = convoluteSrfs(wavelengths,RhowQCGS,RhowStdQC,srfs,inputs['SatelliteSensors'],writer,wb)
#    excelWriter = writer
#    signal  = RhowQCGS
#    dsignal = RhowStdQC
#    sensors = inputs['SatelliteSensors']
    
    lambdaSrfs      = {}
    lambdaSrfsMean  = {}
    RhowLambdaSrfs  = {}
    dRhowLambdaSrfs = {}
    RhowBands       = {}
    dRhowBands      = {}
    
    for sensor in sensors:
        lambdaSrfs     [sensor] = srfs[sensor].index
        lambdaSrfsMean [sensor] = {}
        RhowLambdaSrfs [sensor] = pd.DataFrame(columns=lambdaSrfs[sensor])
        dRhowLambdaSrfs[sensor] = pd.DataFrame(columns=lambdaSrfs[sensor])
        RhowBands      [sensor] = pd.DataFrame(columns=srfs[sensor].columns)
        dRhowBands     [sensor] = pd.DataFrame(columns=srfs[sensor].columns)

        # calculate mean wavelength
        for band in srfs[sensor].columns:
            lambdaSrfsMean[sensor][band] = np.sum(list(lambdaSrfs[sensor])*srfs[sensor][band])/np.sum(srfs[sensor][band])

        # interp Rhows to lambdaSrfs:
        for st in signal.index:
#            if np.all(signal.loc[st].isna()):
#                RhowLambdaSrfs[ sensor].loc[st,:] = np.nan
#                dRhowLambdaSrfs[sensor].loc[st,:] = np.nan
#                RhowLambdaSrfs[ sensor] df.append(pandas.Series(), ignore_index=True)
            RhowLambdaSrfs[ sensor].loc[st,:] = np.interp(lambdaSrfs[sensor], wavelengths,  signal.loc[st])
            dRhowLambdaSrfs[sensor].loc[st,:] = np.interp(lambdaSrfs[sensor], wavelengths, dsignal.loc[st])
            for band in srfs[sensor].columns:
                if wavelengths[0] <= lambdaSrfsMean[sensor][band] <= wavelengths[-1]:
                    S = np.sum(srfs[sensor][band])
                    RhowBands[sensor].loc[ st,band] =         np.sum(np.array(RhowLambdaSrfs[sensor].loc[st,:]*srfs[sensor][band]))/S
                    dRhowBands[sensor].loc[st,band] = np.sqrt(np.sum(np.array((dRhowLambdaSrfs[sensor].loc[st,:]*srfs[sensor][band]/S)**2)))
                else:
                    RhowBands[ sensor].loc[st,band] = np.nan
                    dRhowBands[sensor].loc[st,band] = np.nan

        if excelWriter != 'noExcel':
            sheetname = sensor
            if sheetname in wb.sheetnames:
                wb.remove_sheet(wb.get_sheet_by_name(sheetname))
            RhowBands[sensor].to_excel(excelWriter, index = True, sheet_name=sheetname,startrow = 1)
#            adjustColWidth(wb.get_sheet_by_name(sheetname))
            excelWriter.save(); excelWriter.save()
            excelWriter.close(); excelWriter.close()

            sheetname = sensor + 'Std'
            if sheetname in wb.sheetnames:
                wb.remove_sheet(wb.get_sheet_by_name(sheetname))
            dRhowBands[sensor].to_excel(excelWriter, index = True, sheet_name=sheetname,startrow = 1)
#            adjustColWidth(wb.get_sheet_by_name(sheetname))
            excelWriter.save(); excelWriter.save()
            excelWriter.close(); excelWriter.close()
    return RhowBands,dRhowBands,lambdaSrfsMean


#%% Trios Processing
def triosProcess(campaign0,path0):
#    path0 = '/home/gossn/Dropbox/Documents/inSitu/Database'
#    campaign0 = 'RdP_20170126_Muelle_TURBINET'

    pathRegions   = path0 + '/regions'
    pathSrfs      = path0 + '/SRFs'
    pathScripts   = path0 + '/scripts'
    pathTriosCrit = pathScripts + '/TriOS/Trios_criteria'

    region = campaign0.split('_')[0]
    month  = campaign0.split('_')[1]
    pathCampaign  = pathRegions + '/' + region + '/' +  campaign0
    campaign = region + '_' + month
        
    inputs = {}
    try:
        with open(pathCampaign + '/Trios/triosProcessingInputs') as f0:
            for line in f0:
               (key, val) = line.split()
               inputs[key] = val.split(',')
               if len(inputs[key])==1:
                   inputs[key] = inputs[key][0]
    except:
        print('No TriOS Input file or no Trios measurements for this campaign!')
        return
    
    if os.path.isdir(pathCampaign + '/TriosProcessed/'):
        shutil.rmtree(pathCampaign + '/TriosProcessed/')
    os.mkdir(pathCampaign + '/TriosProcessed/')
    
    figPath = pathCampaign + '/TriosProcessed/Figures/'
    if not os.path.isdir(figPath):
        os.mkdir(figPath)
        os.mkdir(figPath + 'png/')
        os.mkdir(figPath + 'pdf/')
        
    
    
    filePath = pathCampaign + '/TriosProcessed/' + campaign + '_LogTriosOutput'
    try:
        os.remove(filePath)
    except:
        pass
    f = open(filePath,'w')
    
    wavelengths = np.arange(float(inputs['lambdaMin']),float(inputs['lambdaMax'])+1,float(inputs['dLambda']))
    
    print('Log for campaign: ' + campaign                                                                                  + '\n',file=f)
    print('Processing PC in timezone: ' + inputs['deltaUTC'] + ' UTC'                                                      + '\n',file=f)
    print('Trios wavelength vector: '   + inputs['lambdaMin'] + ':' + inputs['dLambda'] + ':' + inputs['lambdaMax'] + 'nm' + '\n',file=f)
    print('Ed tilt threshold set to: '  + inputs['EdTiltThresh'] + ' degrees'                                              + '\n',file=f)
    print('QC check: RhowStd(750)<'  + inputs['QCRhowStd750Thresh']                                                        + '\n',file=f)
    
    if float(inputs['glintCorrectWave']) in wavelengths:
        print('Glint correction applied: Rhow(' + inputs['glintCorrectWave'] + ') subtracted!'                             + '\n',file=f)    
    elif float(inputs['glintCorrectWave']) == -1:
        print('No glint correction applied!'                                                                               + '\n',file=f)
    else:
        raise IOError('Glint correction wavelength out of Trios wavelength vector!')
    
    print('Ed Sensor code: '      + inputs['EdSensor']                                                                     + '\n',file=f)
    print('Lse Sensor code: '     + inputs['LseSensor']                                                                    + '\n',file=f)
    print('Lsky Sensor code: '    + inputs['LskySensor']                                                                   + '\n',file=f)
    print('Ed tilt Sensor code: ' + inputs['TiltSensor']                                                                   + '\n',file=f)
    print('Recalibration factors found in : ' + pathTriosCrit + '/' + inputs['recalFactorFile']                            + '\n',file=f)
    print('Compute reflectances for bands in sensors: ' + ''.join(str(e) + ', ' for e in inputs['SatelliteSensors'])       + '\n',file=f)
    
    # Read SRFs
    srfs = {}
    for sensor in inputs['SatelliteSensors']:
        try:
            srfs[sensor] = pd.read_excel(pathSrfs + '/SpectralResponse_' + sensor + '.xlsx').set_index('lambdas')
            srfs[sensor] = srfs[sensor].fillna(0)
            srfs[sensor][srfs[sensor]<0] = 0
            
        except:
            raise IOError('Missing SRF Excel for sensor ' + sensor + ' in: ' + pathSrfs)
    
    
    # Trios Processing
    
    
    font = {'family' : 'serif',
            'weight' : 'bold',
            'size'   : 24}
    plt.rc('font', **font)
    
    colScans = ['r', 'g', 'b', 'y', 'c']
    markMag = ['-','--']
    
    jumCrit = pd.read_csv(pathTriosCrit + '/Spectral_Jump_criteria.txt')
    outlCrit = pd.read_csv(pathTriosCrit + '/Outlier_criteria.txt')
    
    recalCoef = pd.read_csv(pathTriosCrit + '/' + inputs['recalFactorFile'],sep='\t')
    
    
    mobley99Params = [0.0256, 0.00039, 0.000034]
    
    colNamesWave =  [str(wavelengths[w]) for w in range(len(wavelengths))]
    colNamesRad  = ['DateTime','Latitude','Longitude','IntegrationTime'] + colNamesWave
    colNamesIncl = ['DateTime','Latitude','Longitude','incl_x','incl_y','incl']
    
    #['DateTime'] + [col + '_' + magStr[0] for col in colNamesRad]
    
    
    processStepsRad  = ['raw','simult','specJump','specJumpAll','outlier','okScans']
    processStepsIncl = ['raw','simult','tiltCrit'                        ,'okScans']
    processStepsRho  = [      'simult','QCCriteria'                      ,'okScans']
    
    processStepsRad2Excel = ['simult','specJumpAll','outlier']
    processStepsInc2Excel = ['simult','tiltCrit']
    processStepsLw2Excel  = ['simult']
    processStepsRho2Excel = ['simult','QCCriteria']
    
    mag   = { 'Ed'  : {'type': 'rad' , 'fsuffix': '_SAM_' + inputs['EdSensor']   + '_CALIBRATED_SPECTRUM.mlb'  , \
    'colNames':colNamesRad , 'jumpCrit':float(jumCrit.loc[1,:].values[0]), 'outlCrit':float(outlCrit.loc[1,:].values[0]),\
    'name':'Downwelling Irradiance', 'units': 'mW/m^2/nm'   , 'process2Excel':processStepsRad2Excel},\
    
             'Lse'  : {'type': 'rad' , 'fsuffix': '_SAM_' + inputs['LseSensor']  + '_CALIBRATED_SPECTRUM.mlb'  , \
    'colNames':colNamesRad , 'jumpCrit':float(jumCrit.loc[3,:].values[0]), 'outlCrit':float(outlCrit.loc[3,:].values[0]),\
    'name':'Water-leaving Radiance', 'units': 'mW/m^2/sr/nm', 'process2Excel':processStepsRad2Excel},\
    
             'Lsky' : {'type': 'rad' , 'fsuffix': '_SAM_' + inputs['LskySensor'] + '_CALIBRATED_SPECTRUM.mlb'  , \
    'colNames':colNamesRad , 'jumpCrit':float(jumCrit.loc[5,:].values[0]), 'outlCrit':float(outlCrit.loc[5,:].values[0]),\
    'name':'Sky Radiance'          , 'units': 'mW/m^2/sr/nm', 'process2Excel':processStepsRad2Excel},\
    
             'Tilt' : {'type': 'ang' , 'fsuffix': '_IP_'  + inputs['TiltSensor'] + '_Calibrated_Inclination.mlb', \
    'colNames':colNamesIncl                                                                                             ,\
    'name':'Ed Tilt'       , 'units': 'º'         , 'process2Excel':processStepsInc2Excel},\
                       
             'Lw'   : {'type': 'der'                                                                                          , \
    'colNames':colNamesRad                                                                                              ,\
    'name':'Fresnel-corr. Radiance', 'units': 'mW/m^2/sr/nm', 'process2Excel':processStepsLw2Excel },\
    
             'Rhow' : {'type': 'der'                                                                                          , \
    'colNames':colNamesRad                                                                                              ,\
    'name':'Water reflectance'     , 'units': ''            , 'process2Excel':processStepsRho2Excel}}
    
    miscellaneousCols = ['FresnelTrios','Lsky:Ed(750)','Rhow_std(750)','Rhow_std(750)_QC','CV780 [%]','CV860 [%]','MaxCV400:900 [%]','SZA [º]']
    miscellaneous = pd.DataFrame(columns = miscellaneousCols)
    selectedScansAllSt = {}
    
    for m in mag:
        if   mag[m]['type'] == 'rad':
            for ps in processStepsRad:
                mag[m][ps] = {}
        elif mag[m]['type'] == 'ang':
            for ps in processStepsIncl:
                mag[m][ps] = {}
        elif mag[m]['type'] == 'der':
            for ps in processStepsRho:
                mag[m][ps] = {}

    stationInfo     = pd.read_excel(pathCampaign + '/' + campaign + '.xlsx',sheet_name='stationInfo'   ,skiprows=1)
    radiometryMisc  = pd.read_excel(pathCampaign + '/' + campaign + '.xlsx',sheet_name='radiometryMisc',skiprows=1)

    dfCampaign = pd.concat([stationInfo,radiometryMisc['windSpeed[m/s]']],axis=1).set_index('StationID')
    # Remendar bug openpyxl
    dfCampaign['startTimeUTC'] = pd.to_datetime(dfCampaign['startTimeUTC'],format= '%H:%M:%S' )
    try:
        dfCampaign['timeStampUTC'] = dfCampaign['DateUTC'] + pd.to_timedelta(dfCampaign['startTimeUTC']) + timedelta(hours=70*24*365.25-12)
    except:
        pass
    #

    # Setear como string todas las estaciones con ID numerico: 1 a '1', etc...
    dfCampaign.index = [str(i) for i in list(dfCampaign.index)]
    
    print('\nProcessing campaign ' + campaign)
    
    st0 = 0
    for st in dfCampaign.index:
        st0+=1;
        print('\nProcessing station "' + st + '" ( ' + str(st0) + ' out of ' + str(len(dfCampaign)) + ')', file=f)
        print('\nProcessing station "' + st + '" ( ' + str(st0) + ' out of ' + str(len(dfCampaign)) + ')')
    
        # Read raw data from TriOS
        noDataFlag = False
        for m in [m for m in mag if mag[m]['type'] != 'der']:
            try:
                mag0     = pd.read_table(pathCampaign + '/Trios/' + st            + mag[m]['fsuffix'], sep='\t')
            except:
                print('No TriOS measurement for "' + st + '"', file=f)
                print('No TriOS measurement for "' + st + '"')
                noDataFlag = True
                break
            # Adapting format
            mag0 = mag0.rename(index=str, columns={mag0.columns[0]: 'generic'})
            magDf = pd.DataFrame(columns=mag[m]['colNames'])
            for entry in range(len(mag0)): #ent: entry
                entry0 = mag0.iloc[entry,:].values[0].split()
                entry1 = []
                for element in entry0:
                    try:
                        entry1.append(float(element))
                    except:
                        continue
                if len(entry1) == len(mag[m]['colNames']):
                    magDf = magDf.append(pd.Series(entry1, index=mag[m]['colNames']), ignore_index=True)
            print('# Scans ' + m + ': ' + str(len(magDf)), file=f)
            print('# Scans ' + m + ': ' + str(len(magDf)))
            mag[m]['raw'][st]=magDf
        if noDataFlag:
            miscellaneous.loc[st,:] = ['' for c in range(len(miscellaneousCols))]
            continue
        
        # Recal coefficients
        for m in [m for m in mag if mag[m]['type'] == 'rad']:
            sensor = [col for col in recalCoef.columns if mag[m]['fsuffix'][1:8] in col]
            calcoef = recalCoef[['wavelength'] + sensor].set_index('wavelength')
            for w in calcoef.index:
                mag[m]['raw'][st][str(w)] = mag[m]['raw'][st][str(w)]*calcoef.loc[w,:].values[0]
        
        # Retain simultaneous scans
        timeMag = []
        globalTriggerFlag = True
        for m in mag:
            if mag[m]['type'] != 'der':
                timeMag.append(mag[m]['raw'][st]['DateTime'])
        timeSimult = sorted(list(set(timeMag[0]).intersection(*timeMag))) # intersect all 'DateTime's
        if timeSimult ==  []:
            print('No simultaneous scans!! Take first 4 decimal digits of xlsTime (Likely, non-centralized trigger)', file=f)
            print('No simultaneous scans!! Take first 4 decimal digits of xlsTime (Likely, non-centralized trigger)')
            timeMag = [np.round(10000*timeMag[t])/10000 for t in range(len(timeMag))]
            timeSimult = sorted(list(set(timeMag[0]).intersection(*timeMag))) # intersect all 'DateTime's
            globalTriggerFlag = False
            
        print('# Simultaneous scans: ' + str(len(timeSimult)), file=f)
        for m in [m for m in mag if mag[m]['type'] != 'der']:
                if globalTriggerFlag:
                    mag[m]['simult'][st]=mag[m]['raw'][st][mag[m]['raw'][st]['DateTime'].isin(timeSimult)]
                    badScan = (~mag[m]['raw'][st]['DateTime'].isin(timeSimult)).tolist()
                else:
                    mag[m]['simult'][st]=mag[m]['raw'][st][(np.round(10000*mag[m]['raw'][st]['DateTime'])/10000).isin(timeSimult)]
                    badScan = (~(np.round(10000*mag[m]['raw'][st]['DateTime'])/10000).isin(timeSimult)).tolist()
                mag[m]['simult'][st].index = range(len(timeSimult)) # Re-index
                if any(badScan):
                    print('Non-simultaneous ' + m + ' scans: ' + str([i for i in range(len(badScan)) if badScan[i]==True]), file=f)
                    print('Non-simultaneous ' + m + ' scans: ' + str([i for i in range(len(badScan)) if badScan[i]==True]))
    
        # Spectral jumps
        for m in [m for m in mag if mag[m]['type'] == 'rad']:
            mag[m]['specJump'][st] = pd.DataFrame(columns=colNamesWave)
            for w in range(len(colNamesWave)):
                if float(colNamesWave[w])>=400 and float(colNamesWave[w])<=900:
                    mag[m]['specJump'][st][colNamesWave[w]] =   abs(mag[m]['simult'][st][colNamesWave[w+1]] - mag[m]['simult'][st][colNamesWave[w]]) >= \
                                            mag[m]['jumpCrit']*0.5*(mag[m]['simult'][st][colNamesWave[w+1]] + mag[m]['simult'][st][colNamesWave[w]])
                    SJ = mag[m]['specJump'][st].index[mag[m]['specJump'][st][colNamesWave[w]] == True].tolist()
                    if SJ != []:
                        print('Spectral Jump in ' + m + ' @ ' + colNamesWave[w] + 'nm for scan #:  ' + str(SJ), file=f)
                        print('Spectral Jump in ' + m + ' @ ' + colNamesWave[w] + 'nm for scan #:  ' + str(SJ))
            mag[m]['specJumpAll'][st] = mag[m]['specJump'][st].any(1).rename('specJump')
    
        # Temporal Outliers
        for m in [m for m in mag if mag[m]['type'] == 'rad']:
            simultScans = mag[m]['simult'][st].index
            mag[m]['outlier'][st] = pd.Series(index=simultScans, name='outlier')
            outs = []
            for s in range(1,len(simultScans)):
                condOutScan =    abs(mag[m]['simult'][st].loc[simultScans[s  ],'550.0'] \
                                   - mag[m]['simult'][st].loc[simultScans[s-1],'550.0']) >= \
             mag[m]['outlCrit']*0.5*(mag[m]['simult'][st].loc[simultScans[s  ],'550.0'] \
                                   + mag[m]['simult'][st].loc[simultScans[s-1],'550.0'])
                if condOutScan:
                    outs = outs + [s-1,s]
                outs = sorted(list(set(outs)))
            mag[m]['outlier'][st].loc[outs] = True
            mag[m]['outlier'][st] = mag[m]['outlier'][st].fillna(False)
            if outs != []:
                print('Outlier scans in ' + m + ':  ' + str(outs), file=f)
                print('Outlier scans in ' + m + ':  ' + str(outs))
    
        # Inclination angles
        tilt = mag['Tilt']['simult'][st]['incl']
        mag['Tilt']['tiltCrit'][st] = (tilt>=float(inputs['EdTiltThresh'])).rename('tiltCrit')    
        
        # Quality control over Rhow (considering spectral/scan jumps and high tilt angles)
        mag['Rhow']['QCCriteria'][st] = pd.Series(False, index=simultScans, name='QCCriteria')
        for m in mag:
            if   mag[m]['type'] == 'rad':
                cond0 = (mag[m]['specJumpAll'][st]) | (mag[m]['outlier'][st])
            elif mag[m]['type'] == 'ang':
                cond0 = (mag[m]['tiltCrit'][st])
            mag['Rhow']['QCCriteria'][st] = mag['Rhow']['QCCriteria'][st] | cond0
        mag['Rhow']['QCCriteria'][st][mag['Rhow']['QCCriteria'][st]==False] = 'OK'
        mag['Rhow']['QCCriteria'][st][mag['Rhow']['QCCriteria'][st]==True ] = 'Reject'
        mag['Rhow']['QCCriteria'][st] = mag['Rhow']['QCCriteria'][st].rename('QCCriteria')
    
        # Calculate Lw & Rhow
        w = dfCampaign.loc[st,'windSpeed[m/s]']
        qcFlag = mag['Rhow']['QCCriteria'][st] == 'OK'
        cloudy = mag['Lsky']['simult'][st].loc[qcFlag,'750.0'].mean()/mag['Ed']['simult'][st].loc[qcFlag,'750.0'].mean()>0.05
        if np.isnan(w) or cloudy:
            w = 0
        fresnelMobley = mobley99Params[0] + mobley99Params[1]*(w**1) + mobley99Params[2]*(w**2)
        miscellaneous.loc[st,'FresnelTrios'] = fresnelMobley
        mag['Lw'  ]['simult'][st] =        mag['Lse']['simult'][st]-fresnelMobley*mag['Lsky']['simult'][st]
        mag['Rhow']['simult'][st] = np.pi*(mag['Lw' ]['simult'][st])/(mag['Ed']['simult'][st])
    
        # Add Lat-Lon cols to der quantities, drop integration time.    
        for m in [m for m in mag if mag[m]['type'] == 'der']:
            mag[m]['simult'][st][['DateTime','Latitude','Longitude']] = mag['Ed']['simult'][st][['DateTime','Latitude','Longitude']]
            mag[m]['simult'][st].drop(['IntegrationTime'],axis=1)
        
        # Reformat date, supress lat/lon cols:
        for m in mag:
            mag[m]['simult'][st] = mag[m]['simult'][st].drop(['Latitude', 'Longitude'], axis=1)
            if mag[m]['type'] == 'der':
                mag[m]['simult'][st] = mag[m]['simult'][st].drop('IntegrationTime', axis=1)
            mag[m]['simult'][st].rename(columns={'DateTime':'DateTimeXLS'}, inplace=True)
            dateTime = []
            for ent in range(len(mag[m]['simult'][st])):
                dateTime.append(xldate_to_datetime(mag[m]['simult'][st].loc[ent,'DateTimeXLS']-float(inputs['deltaUTC'])/24))
            mag[m]['simult'][st].insert(0, 'DateTimeUTC', dateTime)
        # Delete rejected scans
        
        for m in mag:
            mag[m]['okScans'][st] = mag[m]['simult'][st].loc[qcFlag,:]
    #        mag[m]['okScans'][st].index = range(len(mag[m]['okScans'][st])) # Re-index
    
        # Select first 5 "OK" scans
        for m in mag:
            if len(mag[m]['okScans'][st])<5:
                if m == 'Ed':
                    print('Less than 5 OK scans', file=f)
                    print('Less than 5 OK scans')
                mag[m]['okScans'][st] = pd.DataFrame(np.nan, index=range(5), columns=mag[m]['okScans'][st].columns)
            else:
                mag[m]['okScans'][st] = mag[m]['okScans'][st].loc[list(mag[m]['okScans'][st].index)[0:5],:]
    
        # Compute Statistics of scans
        selectedScans = pd.DataFrame(columns=colNamesWave)
        for m in [m for m in mag if mag[m]['type'] != 'ang']:
            for scan in range(5):
                selectedScans.loc[m + str(scan),:] = mag[m]['okScans'][st].iloc[scan,:]
            selectedScans.loc[m + 'Mean'  ,:] = mag[m]['okScans'][st].mean()
            selectedScans.loc[m + 'Std'   ,:] = mag[m]['okScans'][st].std()
            selectedScans.loc[m + 'CV'    ,:] = 100*selectedScans.loc[m + 'Std',:]/selectedScans.loc[m + 'Mean',:]
        selectedScansAllSt[st] = selectedScans
    
        # Compute Statistics of scans
        miscellaneous.loc[st,'Lsky:Ed(750)'    ] = selectedScansAllSt[st].loc['LskyMean','750.0']/selectedScansAllSt[st].loc['EdMean','750.0']
        miscellaneous.loc[st,'Rhow_std(750)'   ] = selectedScansAllSt[st].loc['RhowStd' ,'750.0']
        miscellaneous.loc[st,'Rhow_std(750)_QC'] = selectedScansAllSt[st].loc['RhowStd' ,'750.0'] < float(inputs['QCRhowStd750Thresh'])
        miscellaneous.loc[st,'CV780 [%]'       ] = selectedScansAllSt[st].loc['RhowCV'  ,'780.0']
        miscellaneous.loc[st,'CV860 [%]'       ] = selectedScansAllSt[st].loc['RhowCV'  ,'860.0']
        miscellaneous.loc[st,'MaxCV400:900 [%]'] = selectedScansAllSt[st].loc['RhowCV'  ,[str(w) for w in wavelengths if w>=400 and w<=900]].max()
    
    
        # Compute Solar Zenith Angle
        try:
            sza = float(90)-get_altitude(dfCampaign.loc[st,'Lat'],dfCampaign.loc[st,'Lon'],dfCampaign.loc[st,'timeStampUTC'].replace(tzinfo=pytz.UTC),0)
        except:
            sza = np.nan
        miscellaneous.loc[st,'SZA [º]'] = sza
    
        # GRAPHS per station
    
        titStr = campaign + ' ' + st + ' ' + str(dfCampaign.loc[st,'timeStampUTC']) + 'UTC\n' + 'SZA = ' + str(np.round(miscellaneous.loc[st,'SZA [º]']))
        figHandles = {}


        # Graphs.1: Plot simultaneous scans of 'Lse', 'Lsky', 'Ed', 'Rhow' (spot "bad scans")
        fnum = 0
        figHandles[fnum] = plt.figure(figsize=(30.0, 25.0))
        magPlot = ['Lse', 'Lsky', 'Ed', 'Rhow']
        sp=-1
        for m in magPlot:
            sp+=1
            scan = -1
            plt.subplot(221+sp)
            for idx in mag[m]['okScans'][st].index:
                scan+=1
                plt.plot(wavelengths,mag[m]['okScans'][st].loc[idx,[str(w) for w in wavelengths]],colScans[scan],label='Scan ' + str(idx))
            plt.plot(wavelengths,selectedScansAllSt[st].loc[m + 'Mean'  ,:],'--k',label='Mean')
            if sp == 0:
                plt.title(titStr)
            plt.xlabel('Wavelength [nm]')
            plt.ylabel(mag[m]['name'] + '\n' + mag[m]['units'])
        plt.legend() 
        manager = plt.get_current_fig_manager(); manager.window.showMaximized(); plt.tight_layout#; plt.show()
        figHandles[fnum].savefig(figPath + 'png/' + campaign + '_' + st + '_' + '_'.join(magPlot) + '.png')
    
        # Graphs.2: Plot "5 selected scans" + statistics, one plot for each of the following: 'Lse', 'Lsky', 'Ed', 'Rhow'
        for m in magPlot:
            fnum+=1
            figHandles[fnum] = plt.figure(figsize=(30.0, 25.0))
            scan = -1
            for idx in mag[m]['okScans'][st].index:
                scan+=1
                plt.plot(wavelengths,mag[m]['okScans'][st].loc[idx,[str(w) for w in wavelengths]],colScans[scan],label='Scan ' + str(idx))
            plt.plot(wavelengths,selectedScansAllSt[st].loc[m + 'Mean'  ,:],'--k',label='Mean')
            plt.title(titStr)
            plt.xlabel('Wavelength [nm]')
            plt.ylabel(mag[m]['name'] + '\n' + mag[m]['units'])
            plt.legend()
            manager = plt.get_current_fig_manager(); manager.window.showMaximized(); plt.tight_layout#; plt.show()
            figHandles[fnum].savefig(figPath + 'png/' + campaign + '_' + st + '_' + m + '.png')
    
        # Graphs.3: 
        fnum+=1
        figHandles[fnum] = plt.figure(figsize=(30.0, 25.0))
        magPlot = ['Lse','Lw']
        mark = -1
        for m in magPlot:
            mark+=1
            scan = -1
            for idx in mag[m]['okScans'][st].index:
                scan+=1
                plt.plot(wavelengths,mag[m]['okScans'][st].loc[idx,[str(w) for w in wavelengths]],colScans[scan] + markMag[mark],label=m + ' Scan ' + str(idx))
            plt.plot(wavelengths,selectedScansAllSt[st].loc[m + 'Mean'  ,:],'k' + markMag[mark],label=m + ' Mean')
            plt.title(titStr)
            plt.xlabel('Wavelength [nm]')
        plt.ylabel('Radiance' + '\n' + mag[m]['units'])
        plt.legend()
        manager = plt.get_current_fig_manager(); manager.window.showMaximized(); plt.tight_layout#; plt.show()
        figHandles[fnum].savefig(figPath + 'png/' + campaign + '_' + st + '_' + '_'.join(magPlot) + '.png')
    
    
        # Graphs.4
        fnum+=1
        figHandles[fnum] = plt.figure(figsize=(30.0, 25.0))
        magPlot = ['Ed', 'Lsky', 'Lse', 'Rhow']
        wavePlot = [400.0,550.0,700.0,850.0]
        sp=-1
        for m in magPlot:
            w0=-1
            sp+=1
            plt.subplot(221+sp)
            for w in wavePlot:
                w0+=1
                plt.plot(range(len(mag[m]['simult'][st])),mag[m]['simult'][st].loc[:,str(w)],colScans[w0],label=str(w) + ' nm')
                
                if mag[m]['type'] == 'rad':
                
                    outlierScans = mag[m]['simult'][st].loc[mag[m]['outlier'][st],str(w)]
                    plt.plot(outlierScans.index,outlierScans,'k.',label='_nolegend_')
        
                    specJumpAllScans = mag[m]['simult'][st].loc[mag[m]['specJumpAll'][st],str(w)]
                    plt.plot(specJumpAllScans.index,specJumpAllScans,'k+',label='_nolegend_')
    
                    tiltScans = mag[m]['simult'][st].loc[mag['Tilt']['tiltCrit'][st],str(w)]
                    plt.plot(tiltScans.index,tiltScans,'kx',label='_nolegend_')
    
                elif mag[m]['type'] == 'der':
            
                    badScans = mag[m]['simult'][st].loc[~qcFlag,str(w)]
                    plt.plot(badScans.index,badScans,'ko',label='_nolegend_')
            
                plt.plot(mag[m]['okScans'][st].index,mag[m]['okScans'][st][str(w)],'co',label='_nolegend_')
    
            plt.plot(np.nan,np.nan,'co',label='First 5 "OK" scans')
            plt.plot(np.nan,np.nan,'k.',label='Outlier')        
            plt.plot(np.nan,np.nan,'k+',label='Spectral Jump')
            plt.plot(np.nan,np.nan,'kx',label='Hi Tilt')
            plt.plot(np.nan,np.nan,'ko',label='Not passed QC')
    
            if sp == 0:
                plt.title(titStr)
            plt.xlabel('Scan number')
            plt.ylabel(mag[m]['name'] + '\n' + mag[m]['units'])
        plt.legend()
        manager = plt.get_current_fig_manager(); manager.window.showMaximized(); plt.tight_layout#; plt.show()
        figHandles[fnum].savefig(figPath + 'png/' + campaign + '_' + st + '_' + '_'.join(magPlot) + '_ScansAt_' + '_'.join([str(int(w)) for w in wavePlot]) + '.png')
    
        # Graphs.5
        fnum+=1
        figHandles[fnum] = plt.figure(figsize=(30.0, 25.0))
        m = 'Rhow'
        scan = -1
        plt.plot(wavelengths,abs(selectedScansAllSt[st].loc[m + 'CV'  ,[str(w) for w in wavelengths]]),colScans[scan],label='CV(' + m + ')')
        plt.title(titStr)
        plt.xlabel('Wavelength [nm]')
        plt.ylabel(mag[m]['name'] + '\n' + mag[m]['units'])
        plt.ylim(0,50)
        plt.legend()
        manager = plt.get_current_fig_manager(); manager.window.showMaximized(); plt.tight_layout#; plt.show()
        figHandles[fnum].savefig(figPath + 'png/' + campaign + '_' + st + '_' + m + 'CV' + '.png')
    
        # Graphs.6
        fnum+=1
        figHandles[fnum] = plt.figure(figsize=(30.0, 25.0))
        m = 'Rhow'
        scan = -1
        plt.errorbar(wavelengths,selectedScansAllSt[st].loc[m + 'Mean'  ,[str(w) for w in wavelengths]],yerr=selectedScansAllSt[st].loc[m + 'Std'  ,[str(w) for w in wavelengths]],xerr=None,ecolor=colScans[scan],label=mag[m]['name'])
        plt.title(titStr)
        plt.xlabel('Wavelength [nm]')
        plt.xlim(400,900)
        plt.ylabel(mag[m]['name'] + '\n' + mag[m]['units'])
        plt.legend()
        manager = plt.get_current_fig_manager(); manager.window.showMaximized(); plt.tight_layout#; plt.show()
        figHandles[fnum].savefig(figPath + 'png/' + campaign + '_' + st + '_' + m + '_ErrorBars' + '.png')
        
        # Graphs.7
        fnum+=1
        figHandles[fnum] = plt.figure(figsize=(30.0, 25.0))
        m = 'Tilt'
        scan=-1
        scan+=1
        plt.axhline(float(inputs['EdTiltThresh']),label='Tilt critical value')
        
        tiltScans = mag[m]['simult'][st].loc[mag['Tilt']['tiltCrit'][st],'incl']
        plt.plot(tiltScans.index,tiltScans,'kx',label='Hi Tilt')
        
        plt.plot(range(len(mag[m]['simult'][st])),mag[m]['simult'][st].loc[:,'incl'],colScans[scan],label=mag[m]['name'])
        plt.title(titStr)
        plt.xlabel('Scan number')
        plt.ylabel(mag[m]['name'] + '\n' + mag[m]['units'])
        plt.legend()
        manager = plt.get_current_fig_manager(); manager.window.showMaximized(); plt.tight_layout#; plt.show()
        figHandles[fnum].savefig(figPath + 'png/' + campaign + '_' + st + '_' + m + '.png')
        
        pdf = matplotlib.backends.backend_pdf.PdfPages(figPath + 'pdf/' + campaign + '_' + st + '.pdf')
        for fig in figHandles: ## will open an empty extra figure :(
            pdf.savefig(figHandles[fig], bbox_inches='tight')
        pdf.close()
        plt.close('all')
        
        # Write to Excel: Station workbook
        pathXlsxStation = pathCampaign + '/TriosProcessed/' + campaign + '_' + st + '.xlsx'
        wb = openpyxl.Workbook()
        wb.save(pathXlsxStation)
        
        writer = pd.ExcelWriter(pathXlsxStation, engine = 'openpyxl')
        writer.book = wb
#        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        del wb['Sheet']
    
        for m in mag:
            mag2Excel = pd.DataFrame(index=simultScans)
            for exc in mag[m]['process2Excel']:
                mag2Excel = pd.concat([mag2Excel,mag[m][exc][st]],axis=1)
            sheetname = m
            if sheetname in wb.sheetnames:
#                wb.remove_sheet(wb.get_sheet_by_name(sheetname))
                del wb[sheetname]
            mag2Excel.to_excel(writer, index = False, sheet_name=sheetname)
#            adjustColWidth(wb.get_sheet_by_name(m))
#            adjustColWidth(wb[m])
            writer.save()
            writer.close()
        
    
        sheetname = 'selectedScans'
        if sheetname in wb.sheetnames:
            wb.remove_sheet(wb.get_sheet_by_name(sheetname))
        selectedScansAllSt[st].to_excel(writer, index = True, sheet_name=sheetname)
#        adjustColWidth(wb.get_sheet_by_name(sheetname))
        writer.save()
        writer.close()
    
    
    ###############################################################################
    
    for qc750Flag in [[False,''], [True,'_QC_RhowStd750']]:
        
        # Write to Excel: General TriOS workbook
    
        pathXlsx = pathCampaign + '/TriosProcessed/' + campaign + '_Trios' + qc750Flag[1] + '.xlsx'
        
        wb = openpyxl.Workbook()
        wb.save(pathXlsx)
        
        writer = pd.ExcelWriter(pathXlsx, engine = 'openpyxl')
        writer.book = wb
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))    
        
        # Write to Excel:
        sheetname = 'stationInfo'
        if sheetname in wb.sheetnames:
            wb.remove_sheet(wb.get_sheet_by_name(sheetname))
        stationInfo.set_index('StationID')
        stationInfo.to_excel(writer, index = False, sheet_name=sheetname,startrow = 1)
        writer.save(); writer.save()
        writer.close(); writer.close()
        
        # Write to Excel: Trios Miscellaneous Info
        sheetname = 'TriosStats'
        if sheetname in wb.sheetnames:
            wb.remove_sheet(wb.get_sheet_by_name(sheetname))
        miscellaneous.to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)
#        adjustColWidth(wb.get_sheet_by_name(sheetname))
        writer.save(); writer.save()
        writer.close(); writer.close()
        
        # Write to Excel: Statistics
        stats = ['Mean','Std','CV']
        for m in [m for m in mag if m != 'Lw' and m != 'Tilt']:
            for s in stats:
                statsDf = pd.DataFrame(columns=colNamesWave)
                for st in dfCampaign.index:
                    statsDf.loc[st,:] = ['' for w in colNamesWave]
                    try:
                        if (not qc750Flag[0]) or (miscellaneous.loc[st,'Rhow_std(750)_QC']):
                            statsDf.loc[st,:] = selectedScansAllSt[st].loc[m + s  ,:]
                    except:
                        pass
    
                # Save rhoW(QC) data to dataframe
                if qc750Flag[0]:
                    if m == 'Rhow':
                        if s == 'Mean':
                            RhowQCGS = statsDf.replace(to_replace='',value=np.nan)
                            if float(inputs['glintCorrectWave']) != -1:
                                RhowQCGS = RhowQCGS.sub(RhowQCGS['%.1f' % float(inputs['glintCorrectWave'])], axis=0)
                                for wave in wavelengths[wavelengths>float(inputs['glintCorrectWave'])]:
                                    RhowQCGS['%.1f' % float(wave)] = RhowQCGS['%.1f' % float(inputs['glintCorrectWave'])] # mas alla de 900 llevar a 0.
                        elif s == 'Std':
                            RhowStdQC = statsDf.replace(to_replace='',value=np.nan)
    
                sheetname = 'Trios' + m.upper() + s
                if sheetname in wb.sheetnames:
                    wb.remove_sheet(wb.get_sheet_by_name(sheetname))            
                statsDf.to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)
#                adjustColWidth(wb.get_sheet_by_name(sheetname))
                writer.save(); writer.save()
                writer.close(); writer.close()
    
    
    ####################### SRF convolution #######################################
    
    
    # Write to Excel: Satellite bands Sensor
    if qc750Flag[0] and float(inputs['glintCorrectWave']) != -1:
        qc750Flag[1] = qc750Flag[1] + '_GlintSubtract' + inputs['glintCorrectWave']
    pathXlsx = pathCampaign + '/TriosProcessed/' + campaign + '_Trios' + qc750Flag[1] + '_SatSensors.xlsx'
    
    wb = openpyxl.Workbook()
    wb.save(pathXlsx)
    writer = pd.ExcelWriter(pathXlsx, engine = 'openpyxl')
    writer.book = wb
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))   
    
    # Write to Excel: StationInfo
    sheetname = 'stationInfo'
    if sheetname in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheetname))
    stationInfo.set_index('StationID')
    stationInfo.to_excel(writer, index = False, sheet_name=sheetname,startrow = 1)
#    adjustColWidth(wb.get_sheet_by_name(sheetname))
    writer.save(); writer.save()
    writer.close(); writer.close()
    
    # Write to Excel: Trios Miscellaneous Info
    sheetname = 'TriosStats'
    if sheetname in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheetname))
    miscellaneous.to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)
#    adjustColWidth(wb.get_sheet_by_name(sheetname))
    writer.save(); writer.save()
    writer.close(); writer.close()
    
    # Write to Excel: RhowQCGS
    sheetname = 'Rhow'
    if sheetname in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheetname))
    RhowQCGS.to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)
#    adjustColWidth(wb.get_sheet_by_name(sheetname))
    writer.save(); writer.save()
    writer.close(); writer.close()
    
    # Write to Excel: RhowQCGS
    sheetname = 'RhowStd'
    if sheetname in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheetname))
    RhowStdQC.to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)
#    adjustColWidth(wb.get_sheet_by_name(sheetname))
    writer.save(); writer.save()
    writer.close(); writer.close()
    

    print('\n')
    print('\n',file=f)

    print('Computing Rhow for sensors specified in variable SatelliteSensors, @ the processing inputs file')
    print('Computing Rhow for sensors specified in variable SatelliteSensors, @ the processing inputs file',file=f)    

    RhowBands,dRhowBands,lambdaSrfsMean = convoluteSrfs(wavelengths,RhowQCGS,RhowStdQC,srfs,inputs['SatelliteSensors'],writer,wb)

    ################################### GRAPHS All Stations #######################
    # Graphs: General
    fnum = -1
    figHandles = {}
    
#    YlimStats = {'Mean': 0.2,'Std': 0.2,'CV': 100}
    m = 'Rhow'
    waveRed = wavelengths[np.logical_and(wavelengths>400,wavelengths<900)]
    for s in stats:
        fnum+=1
        figHandles[fnum] = plt.figure(figsize=(30.0, 25.0))
        for st in dfCampaign.index:
            try:
                plt.plot(waveRed,selectedScansAllSt[st].loc[m + s  ,[str(w) for w in waveRed]],label=st)
            except:
                continue
        plt.title(campaign + ' ' + mag[m]['name'] + ' (' + s + ') [All Stations]')
        plt.xlabel('Wavelength [nm]')
        plt.xlim(400,900)
        if s == 'CV':
            plt.ylim(0,50)
        plt.legend()
        manager = plt.get_current_fig_manager(); manager.window.showMaximized(); plt.tight_layout#; plt.show()
        figHandles[fnum].savefig(figPath + 'png/' + campaign + '_AllStations_' + m + s + '.png')
    
    pdf = matplotlib.backends.backend_pdf.PdfPages(figPath + 'pdf/' + campaign + '_AllStations' + '.pdf')
    for fig in figHandles: ## will open an empty extra figure :(
        pdf.savefig(figHandles[fig], bbox_inches='tight')
    
    # Graphs: General QC
    for sensor in [''] + list(RhowBands.keys()):
        fnum+=1
        figHandles[fnum] = plt.figure(figsize=(30.0, 25.0))
        for st in dfCampaign.index:
            try:
                if not np.all(np.isnan(list(RhowQCGS.loc[st,:]))):
                    p = plt.plot(wavelengths,RhowQCGS.loc[st,:],label=st)
                    if sensor:
                        lambdaSrfsMean[sensor].pop('PAN',None)
#                        plt.plot(list(lambdaSrfsMean[sensor].values()),RhowBands[sensor].loc[st,RhowBands[sensor].columns != 'PAN'],'.',color=p[0].get_color(), label='_nolegend_')
                        plt.errorbar(list(lambdaSrfsMean[sensor].values()),\
                              RhowBands[sensor].loc[st, RhowBands[sensor].columns != 'PAN'],\
                        yerr=dRhowBands[sensor].loc[st,dRhowBands[sensor].columns != 'PAN'],\
                        color=p[0].get_color(), elinewidth = 3, linewidth = 0, label='_nolegend_')
            except:
                continue
        if not sensor:
            plt.plot(wavelengths,(RhowQCGS).mean(axis=0),'k',label='Mean Spectrum')
        plt.plot(wavelengths,0*np.array(wavelengths),'--k',label='_nolegend_')
        titStr = campaign + '_AllStations_RhoW' + qc750Flag[1] + '_' + sensor
        plt.title(titStr)
        plt.xlabel('Wavelength [nm]')
        plt.xlim(wavelengths[0],wavelengths[-1])
#        plt.ylim(-0.01,0.2)
        plt.legend()
        manager = plt.get_current_fig_manager(); manager.window.showMaximized(); plt.tight_layout#; plt.show()
        figHandles[fnum].savefig(figPath + 'png/' + titStr + '.png')
        
        pdf = matplotlib.backends.backend_pdf.PdfPages(figPath + 'pdf/' + campaign + '_AllStations' + '.pdf')
        for fig in figHandles: ## will open an empty extra figure :(
            pdf.savefig(figHandles[fig], bbox_inches='tight')
    pdf.close()
    plt.close('all')
    
    f = open(filePath,'w')

#%% ASD Processing
def asdProcess(campaign0,path0):

    pathRegions   = path0 + '/regions'
    pathSrfs      = path0 + '/SRFs'

    region = campaign0.split('_')[0]
    month  = campaign0.split('_')[1]
    pathCampaign  = pathRegions + '/' + region + '/' +  campaign0
    campaign = region + '_' + month

    inputs = {}
    

    inputsFloat = ['deltaUTC','EdThresh','RhowItrThresh','RhowStdThresh']
    inputsInt   = ['lambdaMin','lambdaMax','dLambda','nRep','nRepPerEd','lambdasQC','glintCorrectWave']

    try:
        with open(pathCampaign + '/ASD/asdProcessingInputs') as f0:
            for line in f0:
                (key, val) = line.split()
                inputs[key] = val.split(',')
                if key in inputsFloat:
                    inputs[key] = [float(f) for f in inputs[key]]
                elif key in inputsInt:
                    inputs[key] = [int(f)   for f in inputs[key]]
                if len(inputs[key])==1:
                    inputs[key] = inputs[key][0]
    except:
        print('No ASD Input file or no ASD measurements for this campaign!')
        return
    if os.path.isdir(pathCampaign + '/ASDProcessed/'):
        shutil.rmtree(pathCampaign + '/ASDProcessed/')
    os.mkdir(pathCampaign + '/ASDProcessed/')

    figPath = pathCampaign + '/ASDProcessed/Figures/'
    if not os.path.isdir(figPath):
        os.mkdir(figPath)
        os.mkdir(figPath + 'png/')
        os.mkdir(figPath + 'pdf/')

    filePath = pathCampaign + '/ASDProcessed/' + campaign + '_LogASDOutput'
    try:
        os.remove(filePath)
    except:
        pass
    f = open(filePath,'w')
    

    print('Log for campaign: ' + campaign + '\n', file=f)
    print('Processing PC in timezone: ' + str(inputs['deltaUTC']) + ' UTC' + '\n', file=f)
    print('ASD wavelength vector: '   + str(inputs['lambdaMin']) + ':' + str(inputs['dLambda']) + ':' + str(inputs['lambdaMax']) + 'nm' + '\n', file=f)

    for (l,k) in zip(inputs['lambdasQC'],range(len(inputs['lambdasQC']))):
        print('QC check: if delta(Ed(' + str(l) + '))>'  + str(inputs['EdThresh'][k]) + ' take scans corresp to Ed/ Ed(' + str(l) + ') is maximum ', file=f)
        print('QC check: outlier if Rhow(' + str(l) + ')>Q3+3/2.max(ITR,'  + str(inputs['RhowItrThresh'][k]) + ')', file=f)
        print('QC check: RhowStd(' + str(l) + ')<'  + str(inputs['RhowStdThresh'][k]) + '\n', file=f)

    wavelengths = np.arange(inputs['lambdaMin'],inputs['lambdaMax']+1,inputs['dLambda'])    
    if inputs['glintCorrectWave'] in wavelengths:
        print('Glint correction applied: Rhow(' + str(inputs['glintCorrectWave']) + ') subtracted!' + '\n', file=f)
    elif inputs['glintCorrectWave'] == -1:
        print('No glint correction applied!' + '\n', file=f)
    else:
        raise IOError('Glint correction wavelength out of ASD wavelength vector!')




    print('Compute reflectances for bands in sensors: ' + ''.join(str(e) + ', ' for e in inputs['SatelliteSensors']) + '\n')
    # Read SRFs
    srfs = {}
    for sensor in inputs['SatelliteSensors']:
        try:
            srfs[sensor] = pd.read_excel(pathSrfs + '/SpectralResponse_' + sensor + '.xlsx').set_index('lambdas')
            srfs[sensor] = srfs[sensor].fillna(0)
            srfs[sensor][srfs[sensor]<0] = 0
            
        except:
            raise IOError('Missing SRF Excel for sensor ' + sensor + ' in: ' + pathSrfs)
    
    
    # ASD Processing
    
    
    font = {'family' : 'dejavu sans',
            'weight' : 'bold',
            'size'   : 24}
    plt.rc('font', **font)
    
    colScans = ['r', 'g', 'b', 'y', 'c', 'ck', 'rg', 'gb', 'gk']
    markMag = ['-','--']
        
    mobley99Params = [0.0256, 0.00039, 0.000034]
    
    colNamesWave =  [int(wavelengths[w]) for w in range(len(wavelengths))]

    #['DateTime'] + [col + '_' + magStr[0] for col in colNamesRad]


    processStepsRad  = ['raw','sets','Mean','Std','CV']
    processStepsDer  = ['raw','sets','Mean','Std','CV']

    mag   = { 'Ed'  : {'type': 'rad' , 'name':'Downwelling Irradiance', 'units': 'mW/m^2/nm'   , 'process2Excel': ['raw']   },\
             'Lse'  : {'type': 'rad' , 'name':'Water-leaving Radiance', 'units': 'mW/m^2/sr/nm', 'process2Excel': ['raw']   },\
             'Lsky' : {'type': 'rad' , 'name':'Sky Radiance'          , 'units': 'mW/m^2/sr/nm', 'process2Excel': ['raw']   },\
             'Lw'   : {'type': 'der' , 'name':'Fresnel-corr. Radiance', 'units': 'mW/m^2/sr/nm', 'process2Excel': ['raw']             },\
             'Rhow' : {'type': 'der' , 'name':'Water reflectance'     , 'units': ''            , 'process2Excel': ['raw']}}

    QCdEd   = [['deltaEd('  + str(l) + ')','deltaEd('  + str(l) + ')_QC']    for l in inputs['lambdasQC']]
    QCRst   = [['Rhow_std(' + str(l) + ')','Rhow_std(' + str(l) + ')_QC']    for l in inputs['lambdasQC']]

    QCdEd   = [item for sublist in QCdEd for item in sublist]
    QCRst   = [item for sublist in QCRst for item in sublist]   

    miscellaneousCols = ['Fresnel','Lsky:Ed(750)','SZA [º]','CV780 [%]','CV860 [%]','MaxCV400:900 [%]','Rhow_outliers','Ed_outliers'] + QCdEd + QCRst
    miscellaneous = pd.DataFrame(columns = miscellaneousCols)
    EdCond       = {}
    outliersEd   = {}
    outliersRhow = {}

    for m in mag:
        if   mag[m]['type'] == 'rad':
            for ps in processStepsRad:
                mag[m][ps] = {}
        elif mag[m]['type'] == 'der':
            for ps in processStepsDer:
                mag[m][ps] = {}

    stationInfo     = pd.read_excel(pathCampaign + '/' + campaign + '.xlsx',sheet_name='stationInfo'   ,skiprows=1)
    radiometryMisc  = pd.read_excel(pathCampaign + '/' + campaign + '.xlsx',sheet_name='radiometryMisc',skiprows=1)

    dfCampaign = pd.concat([stationInfo,radiometryMisc['windSpeed[m/s]']],axis=1).set_index('StationID')
    # Remendar bug openpyxl
    dfCampaign['startTimeUTC'] = pd.to_datetime(dfCampaign['startTimeUTC'],format= '%H:%M:%S' )
    dfCampaign['timeStampUTC'] = dfCampaign['DateUTC'] + pd.to_timedelta(dfCampaign['startTimeUTC']) + timedelta(hours=70*24*365.25-12)

    # Setear como string todas las estaciones con ID numerico: 1 a '1', etc...
    dfCampaign.index = [str(i) for i in list(dfCampaign.index)]

    fileType = {}
    for m in [m for m in mag if mag[m]['type'] == 'rad']:
        fileType[m + 'Files'  ] = []
    nFilesPerEd = 2*inputs['nRepPerEd']+1
    nFiles      = nFilesPerEd*inputs['nRep']
    if   inputs['fileOrder'] == 'standard':
        for f0 in range(nFiles):
            if    f0 % nFilesPerEd      == 0:
                fileType['EdFiles'  ].append(f0)
            elif (f0 % nFilesPerEd) % 2 == 1:
                fileType['LseFiles' ].append(f0)
            elif (f0 % nFilesPerEd) % 2 == 0:
                fileType['LskyFiles'].append(f0)
    elif inputs['fileOrder'] == 'conae1':
        conae1 = {}
        with open(pathCampaign + '/ASD/asdProcessingInputs_conae1') as f0:
            for line in f0:
               (key, val) = line.split()
               conae1[key] = ast.literal_eval(val)
    elif inputs['fileOrder'] == 'custom':
        nFiles = inputs['EdFiles'  ] +\
                 inputs['LseFiles' ] +\
                 inputs['LskyFiles']
        fileType['EdFiles'  ] = inputs['EdFiles'  ]
        fileType['LseFiles' ] = inputs['LseFiles' ]
        fileType['LskyFiles'] = inputs['LskyFiles']

    print('\nProcessing campaign ' + campaign)

#% Loop per station!
    st0 = 0
    for st in dfCampaign.index:
        st0+=1
        print('\nProcessing station "' + st + '" ( ' + str(st0) + ' out of ' + str(len(dfCampaign)) + ')', file=f)
        print('\nProcessing station "' + st + '" ( ' + str(st0) + ' out of ' + str(len(dfCampaign)) + ')')
    
        # Read raw data from ASD
        noDataFlag = False
        try:
            # Get all file names in the station directory
            fileDir = os.listdir(pathCampaign + '/ASD/' + st)
            # Last two digits of file number
            fileNum = [abs(int(''.join(map(str,[int(s) for s in filename if s.isdigit()])))) % 100 for filename in fileDir]
            # Sort files
            fileDir = [x for _,x in sorted(zip(fileNum,fileDir))]
            fileNum = sorted(fileNum)

        except:
            print('No ASD measurement for "' + st + '"', file=f)
            print('No ASD measurement for "' + st + '"')
            miscellaneous.loc[st,:] = ['' for c in range(len(miscellaneousCols))]
            continue

        for m in mag:
            mag[m]['raw'][st] = pd.DataFrame(columns=colNamesWave)

        f0 = -1; ed = -1; lse = -1; lsky = -1
        for file in fileDir:
            f0+=1
            # BASH command: find . -type f -exec sed -i 's/,/./g' {} +
            scan = pd.read_csv(pathCampaign + '/ASD/' + st + '/' + file, sep = '\t').set_index('Wavelength').T

            if   inputs['fileOrder'] == 'standard':
                fileID = f0
            elif inputs['fileOrder'] == 'conae1':
                fileID = fileNum[f0]
                for m in [m for m in mag if mag[m]['type'] == 'rad']:
                    fileRange = conae1[st][m]
                    fileType[m + 'Files'] = list(range(fileRange[0],fileRange[1]+1))
                    if m == 'Ed':
                        fileType[m + 'Files'] = fileType[m + 'Files'][:inputs['nRep']                    ]
                    else:
                        fileType[m + 'Files'] = fileType[m + 'Files'][:inputs['nRep']*inputs['nRepPerEd']]
            elif inputs['fileOrder'] == 'custom':
                fileID = fileNum[f0]

            if fileID in fileType['EdFiles'  ]:
                ed = ed + 1
                scan.index = [str(ed)]
                mag['Ed'  ]['raw'][st] = pd.concat([mag['Ed'  ]['raw'][st], scan], axis=0)
            if fileID in fileType['LseFiles' ]:
                lse = lse + 1
                scan.index = [''.join(map(str,divmod(lse ,inputs['nRepPerEd'])))]
                mag['Lse' ]['raw'][st] = pd.concat([mag['Lse' ]['raw'][st], scan], axis=0)
            if fileID in fileType['LskyFiles']:
                lsky = lsky + 1
                scan.index = [''.join(map(str,divmod(lsky,inputs['nRepPerEd'])))]
                mag['Lsky']['raw'][st] = pd.concat([mag['Lsky']['raw'][st], scan], axis=0)
#            if f0 +1 == nFiles: # Obviar archivos extra
#                break
        if f0 + 1 < nFiles:
            raise IOError('Number of files is less than %s! Check nRep and nRepPerEd values in input file!' % nFiles)

        # Calculate Lw & Rhow
        w = dfCampaign.loc[st,'windSpeed[m/s]']
        cloudy = mag['Lsky']['raw'][st][750].mean()/mag['Ed']['raw'][st][750].mean()>0.05
        if np.isnan(w) or cloudy:
            w = 0
        fresnelMobley = mobley99Params[0] + mobley99Params[1]*(w**1) + mobley99Params[2]*(w**2)
        miscellaneous.loc[st,'Fresnel'] = fresnelMobley
        for n in range(inputs['nRep']):
            for m in range(inputs['nRepPerEd']):
                sIdx = str(n) + str(m)
                mag['Lw'  ]['raw'][st].loc[sIdx,:] =  mag['Lse']['raw'][st].loc[sIdx,:] - mag['Lsky']['raw'][st].loc[sIdx,  :]*fresnelMobley
                mag['Rhow']['raw'][st].loc[sIdx,:] = (mag['Lw' ]['raw'][st].loc[sIdx,:])/(mag['Ed'  ]['raw'][st].loc[str(n),:])
        
        EdRange = {}
        Edwave  = {}
        for l in inputs['lambdasQC']:
            Edwave [l] = mag['Ed']['raw'][st][l].values
            EdRange[l] = np.max(Edwave[l]) - np.min(Edwave[l])
            
        EdRangeAll = np.any([EdRange[inputs['lambdasQC'][l0]]>inputs['EdThresh'][l0] for l0 in range(len(inputs['lambdasQC']))])
        if EdRangeAll:
            idxEdMax = str(Edwave[inputs['lambdasQC'][0]].argmax())
            fitr = 2.5
        else:
            idxEdMax = ''.join(str(x) for x in range(inputs['nRep']))
            fitr = 1.5
############### METODO 1
#        for m in mag:
#            EdCond[m][st]   = [scan for scan in list(mag[m]['raw'][st].index) if scan[0] in idxEdMax]
#            mag[m]['Mean'][st] = mag[m]['raw'][st].loc[EdCond[m][st],:].mean(axis=0)
#            mag[m]['Std' ][st] = mag[m]['raw'][st].loc[EdCond[m][st],:].std(axis=0)
#            mag[m]['CV'  ][st] = mag[m]['Std'][st]/mag[m]['Mean'][st]*100
#        for m in mag:
#            EdCond[m][st]   = [scan for scan in list(mag[m]['raw'][st].index) if scan[0] in idxEdMax]
#            mag[m]['Mean'][st] = mag[m]['raw'][st].loc[EdCond[m][st],:].mean(axis=0)
#            mag[m]['Std' ][st] = mag[m]['raw'][st].loc[EdCond[m][st],:].std(axis=0)
#            mag[m]['CV'  ][st] = mag[m]['raw'][st]/mag[m]['Mean'][st]*100

############### METODO 2
#        EdCond[st]   = [scan for scan in list(mag['Ed']['raw'][st].index) if scan[0] in idxEdMax]
#        for m in mag:
#            mag[m]['sets'][st] = pd.DataFrame()
#            if m != 'Ed':
#                for s in EdCond[st]:
#                    subset = [col for col in mag[m]['raw'][st].index if col[0] == s]
#                    mag[m]['sets'][st] = mag[m]['sets'][st].append((mag[m]['raw'][st].loc[subset,:]).mean(axis=0),ignore_index=True)
#            elif m == 'Ed':
#                mag[m]['sets'][st] = mag[m]['raw'][st]
#
#        for m in mag:
#            mag[m]['Mean'][st] = mag[m]['sets'][st].mean(axis=0)
#            if len(EdCond[st]) == 3:
#                mag[m]['Std' ][st] = mag[m]['sets'][st].std(axis=0)
#            elif len(EdCond[st]) == 1:
#                subset = [col for col in mag[m]['raw'][st].index if col[0] == EdCond[st][0]]
#                if inputs['nRepPerEd'] == 1:
#                    mag[m]['Std' ][st] = (mag[m]['raw'][st].loc[subset,:]).multiply(other = 0).iloc[0]
#                else:
#                    mag[m]['Std' ][st] = (mag[m]['raw'][st].loc[subset,:]).std(axis=0)
#            mag[m]['CV'  ][st] = mag[m]['Std' ][st]/mag[m]['Mean'][st]*100

############### METODO 3

#caca = np.array([-1, 0, 0.1, 3, 15, 17.2, 18, 18.4, 20])
#
#print(T1)
#print(T2)
#print(ITR)

#for k in range(21):
#    caca = np.array([0, k, 20])
#    T1 = np.percentile(caca,3/8*100,interpolation='linear')
#    T2 = np.percentile(caca,5/8*100,interpolation='linear')
#    ITR = T2-T1
#    print('k = ' + str(k))
#    print(T1 - fitr * ITR)
#    print(T2 + fitr * ITR)
#
#    for c in caca:
#        if (c < (T1 - fitr * ITR)) or (c > (T2 + fitr * ITR)):
#            print(str(c) + ' is outlier')
#
#for c in range(len(caca)):
#    print(np.percentile(caca,c/(len(caca)-1)*100,axis=0,interpolation='linear'))

        EdCond[st]   = [scan for scan in list(mag['Ed']['raw'][st].index) if scan[0] in idxEdMax]
        for m in mag:
            mag[m]['sets'][st] = pd.DataFrame()
            if m != 'Ed':
                for s in EdCond[st]:
                    subset          = [col for col in mag[m]['raw'][st].index if col[0] == s]
                    mag[m]['sets'][st] = mag[m]['sets'][st].append(mag[m]['raw'][st].loc[subset,:],ignore_index=False)
                outliersEd[st]  = [col for col in mag[m]['raw'][st].index if col not in mag[m]['sets'][st].index]
            elif m == 'Ed':
                mag[m]['sets'][st] = mag[m]['raw'][st]
        for m in mag:
            if m == 'Rhow':
                outliers = []
                rangeQC = inputs['lambdasQC']
                T1 = np.percentile(mag[m]['sets'][st].loc[:,rangeQC],3/8*100,axis=0,interpolation='linear')
                T2 = np.percentile(mag[m]['sets'][st].loc[:,rangeQC],5/8*100,axis=0,interpolation='linear')
                ITR = np.maximum(T2-T1,np.array(inputs['RhowItrThresh']))
                # En vez de IQR (inter quartile range) uso el ITR (inter tercile range) porque son 9 espectros generalmente... Si se nubla,
                # es probable que se arruinen los 3 espectros asociados a determinada repeticion de la serie. Entonces, los tres serian outliers
                # y quedarian por encima/debajo del tercil, mientras que no ocurriria lo mismo en el caso de los cuartiles
                # El factor 9/4 proviene de asumir que a IQR se le aplica el factor 3/2... entonces, si pienso el IQR como el 50% del rango,
                # y al ITR como el 33% del rango, 3/2*IQR ~ 9/4*ITR... (CREO QUE LO VOY A DEJAR EN 3/2)
                if len(rangeQC) == 1:
                    outliers = np.array((mag[m]['sets'][st].loc[:,rangeQC] < (T1 - fitr * ITR)) | \
                                        (mag[m]['sets'][st].loc[:,rangeQC] > (T2 + fitr * ITR))       )
                    outliers = [out[0] for out in outliers]
                else:
                    outliers = np.any(  (mag[m]['sets'][st].loc[:,rangeQC] < (T1 - fitr * ITR)) | \
                                        (mag[m]['sets'][st].loc[:,rangeQC] > (T2 + fitr * ITR)),axis=1)
                mag[m]['sets'][st].loc[outliers,:] = np.nan
                outliersRhow[st] = list(mag[m]['sets'][st].index[outliers])
            mag[m]['Mean'][st] = mag[m]['sets'][st].mean(axis=0)
            mag[m]['Std' ][st] = mag[m]['sets'][st].std(axis=0)
            mag[m]['CV'  ][st] = mag[m]['Std' ][st]/mag[m]['Mean'][st]*100


        for m in mag:
            for t in ['raw','Mean','Std','CV']:
                mag[m][t][st].index.name = 'Wavelength[nm]'

        # Compute Statistics of scans
        for [l,l0] in zip(inputs['lambdasQC'],range(len(inputs['lambdasQC']))):
            miscellaneous.loc[st,'deltaEd(' + str(l) + ')'    ] = EdRange[l]
            miscellaneous.loc[st,'deltaEd(' + str(l) + ')_QC' ] = EdRange[l] < inputs['EdThresh'][l0]
            miscellaneous.loc[st,'Rhow_std(' + str(l) + ')'   ] = mag['Rhow']['Std' ][st][l]
            miscellaneous.loc[st,'Rhow_std(' + str(l) + ')_QC'] = mag['Rhow']['Std' ][st][l] < inputs['RhowStdThresh'][l0]
        miscellaneous.index.name = 'StationID'
        miscellaneous.loc[st,'Lsky:Ed(750)'    ] = mag['Lsky']['Mean'][st][750]/mag['Ed'  ]['Mean'][st][750]
        miscellaneous.loc[st,'Rhow_outliers'   ] = ','.join(outliersRhow[st])
        miscellaneous.loc[st,'Ed_outliers'     ] = ','.join(outliersEd  [st])
        miscellaneous.loc[st,'CV780 [%]'       ] = mag['Rhow']['CV'  ][st][780]
        miscellaneous.loc[st,'CV860 [%]'       ] = mag['Rhow']['CV'  ][st][860]
        miscellaneous.loc[st,'MaxCV400:900 [%]'] = mag['Rhow']['CV'  ][st][[w for w in wavelengths if w>=400 and w<=900]].max()
        # Compute Solar Zenith Angle
        try:
            sza = float(90)-get_altitude(dfCampaign.loc[st,'Lat'],dfCampaign.loc[st,'Lon'],dfCampaign.loc[st,'timeStampUTC'].replace(tzinfo=pytz.UTC),0)
        except:
            sza = np.nan
        miscellaneous.loc[st,'SZA [º]'] = sza

        # GRAPHS per station

        titStr = campaign + ' ' + st + ' ' + str(dfCampaign.loc[st,'timeStampUTC']) + 'UTC\n' + 'SZA = ' + str(np.round(miscellaneous.loc[st,'SZA [º]']))
        figHandles = {}

        fnum = -1
        
        waveReduced = [w for w in wavelengths if w<inputs['glintCorrectWave']]

        # Graphs.1: Plot simultaneous scans of 'Lse', 'Lsky', 'Ed', 'Rhow' (spot "bad scans")
        fnum+=1
        figHandles[fnum] = plt.figure(figsize=(30.0, 25.0))
        magPlot = ['Ed', 'Lse', 'Lsky', 'Rhow']
        sp=-1
        for m in magPlot:
            sp+=1
            plt.subplot(221+sp)
            for idx in mag[m]['raw'][st].index:
                labelScan = 'Scan ' + str(idx)
                whiteLine = False
                if m =='Rhow':
                    if idx[0] not in EdCond[st]:
                        labelScan = labelScan + ' (bad Ed) '
                        whiteLine = True
                    if idx in outliersRhow[st]:
                        labelScan = labelScan + ' (out) '
                        whiteLine = True
                plt.plot(waveReduced,mag[m]['raw'][st].loc[idx,waveReduced],'-',label=labelScan)
                if whiteLine:
                    plt.plot(waveReduced,mag[m]['raw'][st].loc[idx,waveReduced],'--w',label='_no_legend_')
            plt.plot(waveReduced,mag[m]['Mean'][st][waveReduced],'--k',label='Mean')
            plt.legend()
            plt.xlim(waveReduced[0],waveReduced[-1])
            if sp == 0:
                plt.title(titStr)
            plt.xlabel('Wavelength [nm]')
            plt.ylabel(mag[m]['name'] + '\n' + mag[m]['units'])

        manager = plt.get_current_fig_manager(); manager.window.showMaximized(); plt.tight_layout#; plt.show()
        figHandles[fnum].savefig(figPath + 'png/' + campaign + '_' + st + '_' + '_'.join(magPlot) + '.png')

        # Graphs.2: Lsky/Ed
        fnum+=1
        magPlot = 'Lsky/Ed'
        figHandles[fnum] = plt.figure(figsize=(30.0, 25.0))
        plt.plot(waveReduced,mag['Lsky']['Mean'][st][waveReduced]/mag['Ed'  ]['Mean'][st][waveReduced],label='Lsky/Ed')
        plt.legend()
        plt.xlim(waveReduced[0],waveReduced[-1])
        plt.title(titStr)
        plt.xlabel('Wavelength [nm]')
        plt.ylabel(magPlot)

        manager = plt.get_current_fig_manager(); manager.window.showMaximized(); plt.tight_layout#; plt.show()
        figHandles[fnum].savefig(figPath + 'png/' + campaign + '_' + st + '_Lsky_O_Ed' + '.png')
    

        # Graphs.3: rhoMean
        fnum+=1
        magPlot = 'Rhow'
        figHandles[fnum] = plt.figure(figsize=(30.0, 25.0))
        plt.plot(waveReduced,mag['Rhow']['Mean'][st][waveReduced],label='Mean Rhow')
        plt.legend()
        plt.xlim(waveReduced[0],waveReduced[-1])
        plt.title(titStr)
        plt.xlabel('Wavelength [nm]')
        plt.ylabel(magPlot)

        manager = plt.get_current_fig_manager(); manager.window.showMaximized(); plt.tight_layout#; plt.show()
        figHandles[fnum].savefig(figPath + 'png/' + campaign + '_' + st + '_' + '_RhowMean' + '.png')


        # Graphs.4: rhoCV
        fnum+=1
        magPlot = 'RhowCV'
        figHandles[fnum] = plt.figure(figsize=(30.0, 25.0))
        plt.plot(waveReduced,mag['Rhow']['CV'][st][waveReduced],label='RhowCV')
        plt.legend()
        plt.xlim(waveReduced[0],waveReduced[-1])
        plt.ylim(0,100)
        plt.title(titStr)
        plt.xlabel('Wavelength [nm]')
        plt.ylabel(magPlot)

        manager = plt.get_current_fig_manager(); manager.window.showMaximized(); plt.tight_layout#; plt.show()
        figHandles[fnum].savefig(figPath + 'png/' + campaign + '_' + st + '_' + '_RhowCV' + '.png')

        # Graphs.5: Error Bars
        fnum+=1
        figHandles[fnum] = plt.figure(figsize=(30.0, 25.0))
        m = 'Rhow'
        plt.errorbar(waveReduced,mag['Rhow']['Mean'][st][waveReduced],yerr=mag['Rhow']['Std'][st][waveReduced],label=mag[m]['name'])
        plt.title(titStr)
        plt.xlabel('Wavelength [nm]')
        plt.xlim(waveReduced[0],waveReduced[-1])
        plt.ylabel(mag[m]['name'] + '\n' + mag[m]['units'])
        plt.legend()
        manager = plt.get_current_fig_manager(); manager.window.showMaximized(); plt.tight_layout#; plt.show()
        figHandles[fnum].savefig(figPath + 'png/' + campaign + '_' + st + '_' + m + '_ErrorBars' + '.png')
        
        pdf = matplotlib.backends.backend_pdf.PdfPages(figPath + 'pdf/' + campaign + '_' + st + '.pdf')
        for fig in figHandles: ## will open an empty extra figure :(
            pdf.savefig(figHandles[fig], bbox_inches='tight')
        pdf.close()
        plt.close('all')

        # Write to Excel: Station workbook
        pathXlsxStation = pathCampaign + '/ASDProcessed/' + campaign + '_' + st + '.xlsx'
        wb = openpyxl.Workbook()
        wb.save(pathXlsxStation)
        
        writer = pd.ExcelWriter(pathXlsxStation, engine = 'openpyxl')
        writer.book = wb
#        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        del wb['Sheet']

        for m in mag:
            mag[m]['process2Excel'] = ['raw']
            for exc in mag[m]['process2Excel']:
                mag2Excel = mag[m][exc][st].T
                sheetname = m
                if sheetname in wb.sheetnames:
    #                wb.remove_sheet(wb.get_sheet_by_name(sheetname))
                    del wb[sheetname]
                mag2Excel.to_excel(writer, index = True, sheet_name=sheetname)
#                adjustColWidth(wb.get_sheet_by_name(m))
#                adjustColWidth(wb[m])
                writer.save()
                writer.close()

        for t in ['Mean','Std','CV']:
            sheetname = t
            if sheetname in wb.sheetnames:
                wb.remove_sheet(wb.get_sheet_by_name(sheetname))
            mag2Excel = pd.DataFrame()
            for m in mag:
                mag2Excel[m]= mag[m][t][st].T
            mag2Excel.index_name = 'Wavelength[nm]'
            mag2Excel.to_excel(writer, index = True, sheet_name=sheetname)
#            adjustColWidth(wb.get_sheet_by_name(sheetname))
            writer.save()
            writer.close()

################################################################################
    for qcFlag in [[False,''], [True,'_QC_RhowStd']]:
        
        # Write to Excel: General ASD workbook
    
        pathXlsx = pathCampaign + '/ASDProcessed/' + campaign + '_ASD' + qcFlag[1] + '.xlsx'
        
        wb = openpyxl.Workbook()
        wb.save(pathXlsx)
        
        writer = pd.ExcelWriter(pathXlsx, engine = 'openpyxl')
        writer.book = wb
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        
        
        # Write to Excel:
        sheetname = 'stationInfo'
        if sheetname in wb.sheetnames:
            wb.remove_sheet(wb.get_sheet_by_name(sheetname))
        stationInfo.set_index('StationID')
        stationInfo.to_excel(writer, index = False, sheet_name=sheetname,startrow = 1)
        writer.save(); writer.save()
        writer.close(); writer.close()
        
        # Write to Excel: ASD Miscellaneous Info
        sheetname = 'ASDStats'
        if sheetname in wb.sheetnames:
            wb.remove_sheet(wb.get_sheet_by_name(sheetname))
        miscellaneous.to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)
#        adjustColWidth(wb.get_sheet_by_name(sheetname))
        writer.save(); writer.save()
        writer.close(); writer.close()
        
        # Write to Excel: Statistics
        stats = ['Mean','Std']
        for m in [m for m in mag if m != 'Lw']:
            for s in stats:
                statsDf = pd.DataFrame(columns=colNamesWave)
                for st in dfCampaign.index:
                    statsDf.loc[st,:] = ['' for w in colNamesWave]
                    try:
                        qcStd = True
                        for l in inputs['lambdasQC']:
                            qcStd = qcStd and (miscellaneous.loc[st,'Rhow_std(' + str(l) + ')_QC'])
                        if (not qcFlag[0]) or qcStd:
                            statsDf.loc[st,:] = mag[m][s][st]
                    except:
                        pass
    
                # Save rhoW(QC) data to dataframe
                if qcFlag[0]:
                    if m == 'Rhow':
                        if s == 'Mean':
                            RhowQCGS = statsDf.replace(to_replace='',value=np.nan)
                            if inputs['glintCorrectWave'] != -1:
                                RhowQCGS = RhowQCGS.sub(RhowQCGS[inputs['glintCorrectWave']], axis=0)
                                for wave in wavelengths[wavelengths>inputs['glintCorrectWave']]:
                                    RhowQCGS[wave] = RhowQCGS[inputs['glintCorrectWave']] # mas alla de 900 llevar a 0.
                        elif s == 'Std':
                            RhowStdQC = statsDf.replace(to_replace='',value=np.nan)
    
                sheetname = 'ASD' + m.upper() + s
                if sheetname in wb.sheetnames:
                    wb.remove_sheet(wb.get_sheet_by_name(sheetname))
                statsDf = statsDf.T
                statsDf.to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)
#                adjustColWidth(wb.get_sheet_by_name(sheetname))
                writer.save(); writer.save()
                writer.close(); writer.close()

    ####################### SRF convolution #######################################
    
    
    # Write to Excel: Satellite bands Sensor
    if qcFlag[0] and inputs['glintCorrectWave'] != -1:
        qcFlag[1] = qcFlag[1] + '_GlintSubtract' + str(inputs['glintCorrectWave'])
    pathXlsx = pathCampaign + '/ASDProcessed/' + campaign + '_ASD' + qcFlag[1] + '_SatSensors.xlsx'
    
    wb = openpyxl.Workbook()
    wb.save(pathXlsx)
    writer = pd.ExcelWriter(pathXlsx, engine = 'openpyxl')
    writer.book = wb
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))   
   
    # Write to Excel: StationInfo
    sheetname = 'stationInfo'
    if sheetname in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheetname))
    stationInfo.set_index('StationID')
    stationInfo.to_excel(writer, index = False, sheet_name=sheetname,startrow = 1)
#    adjustColWidth(wb.get_sheet_by_name(sheetname))
    writer.save(); writer.save()
    writer.close(); writer.close()

    # Write to Excel: ASD Miscellaneous Info
    sheetname = 'ASDStats'
    if sheetname in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheetname))
    miscellaneous.to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)
#    adjustColWidth(wb.get_sheet_by_name(sheetname))
    writer.save(); writer.save()
    writer.close(); writer.close()
    
    # Write to Excel: RhowQCGS
    sheetname = 'Rhow'
    if sheetname in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheetname))
    (RhowQCGS.T).to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)
#    adjustColWidth(wb.get_sheet_by_name(sheetname))
    writer.save(); writer.save()
    writer.close(); writer.close()

    # Write to Excel: RhowQCGS
    sheetname = 'RhowStd'
    if sheetname in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheetname))
    (RhowStdQC.T).to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)
#    adjustColWidth(wb.get_sheet_by_name(sheetname))
    writer.save(); writer.save()
    writer.close(); writer.close()

    print('\n')
    print('\n',file=f)
    print('Computing Rhow for sensors specified in variable SatelliteSensors, @ the processing inputs file')
    print('Computing Rhow for sensors specified in variable SatelliteSensors, @ the processing inputs file',file=f)
    
    RhowBands,dRhowBands,lambdaSrfsMean = convoluteSrfs(wavelengths,RhowQCGS,RhowStdQC,srfs,inputs['SatelliteSensors'],writer,wb)
#%
    ################################### GRAPHS All Stations #######################
    # Graphs: General
    fnum = -1
    figHandles = {}

#    YlimStats = {'Mean': 0.2,'Std': 0.2,'CV': 100}
    m = 'Rhow'
    for s in stats:
        fnum+=1
        figHandles[fnum] = plt.figure(figsize=(30.0, 25.0))
        for st in dfCampaign.index:
            try:
                plt.plot(waveReduced,mag[m][s][st][waveReduced],label=st)
            except:
                continue
        plt.title(campaign + ' ' + mag[m]['name'] + ' (' + s + ') [All Stations]')
        plt.xlabel('Wavelength [nm]')
        plt.xlim(waveReduced[0],waveReduced[-1])
        if s == 'CV':
            plt.ylim(0,100)
        plt.legend()
        manager = plt.get_current_fig_manager(); manager.window.showMaximized(); plt.tight_layout#; plt.show()
        figHandles[fnum].savefig(figPath + 'png/' + campaign + '_AllStations_' + m + s + '.png')
    
    pdf = matplotlib.backends.backend_pdf.PdfPages(figPath + 'pdf/' + campaign + '_AllStations' + '.pdf')
    for fig in figHandles: ## will open an empty extra figure :(
        pdf.savefig(figHandles[fig], bbox_inches='tight')
    
    # Graphs: General QC
    for sensor in [''] + list(RhowBands.keys()):
        fnum+=1
        figHandles[fnum] = plt.figure(figsize=(30.0, 25.0))
        for st in dfCampaign.index:
            try:
                if not np.all(np.isnan(list(RhowQCGS.loc[st,:]))):
                    p = plt.plot(waveReduced,RhowQCGS.loc[st,waveReduced],label=st)
                    if sensor:
                        lambdaSrfsMean[sensor].pop('PAN',None)
#                        plt.plot(list(lambdaSrfsMean[sensor].values()),RhowBands[sensor].loc[st,RhowBands[sensor].columns != 'PAN'],'.',color=p[0].get_color(), label='_nolegend_')
                        plt.errorbar(list(lambdaSrfsMean[sensor].values()),\
                              RhowBands[sensor].loc[st, RhowBands[sensor].columns != 'PAN'],\
                        yerr=dRhowBands[sensor].loc[st,dRhowBands[sensor].columns != 'PAN'],\
                        color=p[0].get_color(), elinewidth = 3, linewidth = 0, label='_nolegend_')
            except:
                continue
        if not sensor:
            plt.plot(waveReduced,(RhowQCGS[waveReduced]).mean(axis=0),'k',label='Mean Spectrum')
        plt.plot(wavelengths,0*np.array(wavelengths),'--k',label='_nolegend_')
        titStr = campaign + '_AllStations_RhoW' + qcFlag[1] + '_' + sensor
        plt.title(titStr)
        plt.xlabel('Wavelength [nm]')
        plt.xlim(waveReduced[0],waveReduced[-1])
#        plt.ylim(-0.01,0.2)
        plt.legend()
        manager = plt.get_current_fig_manager(); manager.window.showMaximized(); plt.tight_layout#; plt.show()
        figHandles[fnum].savefig(figPath + 'png/' + titStr + '.png')
        
        pdf = matplotlib.backends.backend_pdf.PdfPages(figPath + 'pdf/' + campaign + '_AllStations' + '.pdf')
        for fig in figHandles: ## will open an empty extra figure :(
            pdf.savefig(figHandles[fig], bbox_inches='tight')
    pdf.close()
    plt.close('all')
    
    f = open(filePath,'w')

#%% TDogliotti & SPMNechad
def TDogliottiSPMNechad(campaign0,path0):

#    path0 = '/home/gossn/Dropbox/Documents/inSitu/Database'
#    campaign0 = 'RdP_20181202_Houssay'

    pathRegions   = path0 + '/regions'
    pathSrfs      = path0 + '/SRFs'

    region = campaign0.split('_')[0]
    month  = campaign0.split('_')[1]
    pathCampaign  = pathRegions + '/' + region + '/' +  campaign0
    campaign = region + '_' + month

    for algo in ['TDogliotti','SPMNechad']:
        print(algo + ' for campaign ' + campaign + '!')
        # Valores del algoritmo de SPM de Nechad (segundos valores: espurios)
        if algo == 'TDogliotti':
            A = [228.1 ,3078.9]
            B = [0     ,0     ]
            C = [0.1641,0.2112]
        if algo == 'SPMNechad':
            A = [253.51,0     ]
            B = [2.32  ,0     ]
            C = [0.1641,1     ]
        
        sensors = {'Hyper': {'bands': [645,860],'A': A, 'B': B, 'C': C},\
                   'MA'   : {'bands': [645,859],'A': A, 'B': B, 'C': C},\
                   'MT'   : {'bands': [645,859],'A': A, 'B': B, 'C': C},\
                   'ME'   : {'bands': [664,865],'A': A, 'B': B, 'C': C},\
                   'MSI'  : {'bands': [665,865],'A': A, 'B': B, 'C': C},\
                   'OLI'  : {'bands': [655,865],'A': A, 'B': B, 'C': C},\
                   'OLCI' : {'bands': [665,865],'A': A, 'B': B, 'C': C},\
                   'PHR1B': {'bands': [656,853],'A': A, 'B': B, 'C': C},\
                   'VIIRS': {'bands': [671,862],'A': A, 'B': B, 'C': C}}
        
        Algo = pd.DataFrame()
    
        for rad in ['Trios','ASD']:
            try:
                pathRhows = glob.glob(pathCampaign + '/' + rad + 'Processed/' + campaign + '*SatSensors.xlsx')[0]
            except:
                print('NO ' + rad + ' measurements for this campaign!')
                continue
            for sensor in sensors.keys():
                T = {}
                dT = {}
                sheetname = sensor
                if sensor == 'Hyper':
                    sheetname = 'Rhow'
                try:
                    rhows  = pd.read_excel(pathRhows,sheet_name=sheetname        ,skiprows=1,index=True)
                    drhows = pd.read_excel(pathRhows,sheet_name=sheetname + 'Std',skiprows=1,index=True)
                except:
                    print('NO rhoW values for ' + sensor + ' were computed for this campaign!')
                    continue
                if sensor == 'Hyper' and rad == 'ASD':
                    rhows  = rhows.T
                    rhows.columns = rhows.iloc[0]
                    rhows  = rhows.iloc[1:,:]
                    rhows.index.name = 'StationID'
                    drhows = drhows.T
                    drhows.columns = drhows.iloc[0]
                    drhows = drhows.iloc[1:,:]
                    drhows.index.name = 'StationID'
                else:
                    rhows  = rhows.rename(columns={'Unnamed: 0': 'StationID'})
                    rhows.set_index('StationID',inplace=True)
                    drhows = drhows.rename(columns={'Unnamed: 0': 'StationID'})
                    drhows.set_index('StationID',inplace=True)
                
                if rad == 'Trios' and sensor == 'Hyper':
                    cols = [str(sensors[sensor]['bands'][j]) + '.0' for j in [0,1]]
                else:
                    cols = [    sensors[sensor]['bands'][j]         for j in [0,1]]

                Algo['RAD'] = rad
                Algo.loc[np.isnan(rhows[cols[0]]),'RAD'] = np.nan
                for bd in [0,1]:
                    T[bd]  = ( rhows[cols[bd]]*sensors[sensor]['A'][bd])/(1-(rhows[cols[bd]]/sensors[sensor]['C'][bd])) + sensors[sensor]['B'][bd]
                    dT[bd] = (drhows[cols[bd]]*sensors[sensor]['A'][bd])/(1-(rhows[cols[bd]]/sensors[sensor]['C'][bd]))**2                    
                if algo == 'TDogliotti':
                    red2nir  = (rhows[cols[0]]-0.05)/0.02
                    red2nir  = np.maximum(np.minimum(red2nir,1),0)
                    dred2nir = drhows[cols[0]]/0.02
                    dred2nir.loc[(red2nir==0) | (red2nir==1)] = 0
                    if sensor == 'Hyper':
                        Algo.loc[:,'Weight[0:Red,1:NIR]'   ] = red2nir
                        Algo.loc[:,'Weight[0:Red,1:NIR]_CV'] = dred2nir/red2nir*100
                if algo == 'SPMNechad':
                    red2nir = 0
                Algo.loc[:,sensor] = T[0]*(1-red2nir) + T[1]*red2nir
                Algo.loc[:,sensor + '_Std'] = dT[0]*(1-red2nir) + dT[1]*red2nir
                Algo.loc[:,sensor + '_CV' ] = Algo.loc[:,sensor + '_Std']/Algo.loc[:,sensor]*100
    
            pathXlsx = pathCampaign + '/' + campaign + '_' + algo + '.xlsx'
            
            wb = openpyxl.Workbook()
            wb.save(pathXlsx)
            
            writer = pd.ExcelWriter(pathXlsx, engine = 'openpyxl')
            writer.book = wb
            wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
            
            
            # Write to Excel:
            sheetname = algo
            if sheetname in wb.sheetnames:
                wb.remove_sheet(wb.get_sheet_by_name(sheetname))
            Algo.to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)
            writer.save(); writer.save()
            writer.close(); writer.close()
    return
#%% CHL, Gitelson 2008 + RDI, Shen 2019
def ThreeBandModels(campaign0,path0):

#    path0 = '/home/gossn/Dropbox/Documents/inSitu/Database'
#    campaign0 = 'RdP_20181202_Houssay'

    pathRegions   = path0 + '/regions'
    pathSrfs      = path0 + '/SRFs'

    region = campaign0.split('_')[0]
    month  = campaign0.split('_')[1]
    pathCampaign  = pathRegions + '/' + region + '/' +  campaign0
    campaign = region + '_' + month

    sensors0 = {}
    
    sensors0['CHL_Gitelson2008'] = \
              {'Hyper': {'bands': [[660,670],[700,730],[740,760]]},\
               'ME'   : {'bands': [664      ,708      ,753      ]},\
               'MSI'  : {'bands': [665      ,705      ,740      ]},\
               'OLCI' : {'bands': [673.75   ,708.75   ,753.75   ]}}

    sensors0['RDI_Shen2019'    ] = \
              {'Hyper': {'bands': [[660,670],[555,565],[740,760]]},\
               'MA'   : {'bands': [667      ,555      ,748      ]},\
               'MT'   : {'bands': [667      ,555      ,748      ]},\
               'ME'   : {'bands': [664      ,559      ,753      ]},\
               'MSI'  : {'bands': [665      ,560      ,740      ]},\
               'OLCI' : {'bands': [665      ,560      ,753.75   ]},\
               'VIIRS': {'bands': [671      ,551      ,745      ]}}

    for algo in ['CHL_Gitelson2008','RDI_Shen2019']:
        sensors = sensors0[algo]
        print('\n' + algo + ' for campaign ' + campaign + '!')
        # Valores del algoritmo de SPM de Nechad (segundos valores: espurios)
                
        Algo = pd.DataFrame()
    
        for rad in ['Trios','ASD']:
            try:
                pathRhows = glob.glob(pathCampaign + '/' + rad + 'Processed/' + campaign + '*SatSensors.xlsx')[0]
            except:
                print('NO ' + rad + ' measurements for this campaign!')
                continue
            for sensor in sensors.keys():
                T = {}
                dT = {}
                sheetname = sensor
                if sensor == 'Hyper':
                    sheetname = 'Rhow'
                try:
                    rhows  = pd.read_excel(pathRhows,sheet_name=sheetname        ,skiprows=1,index=True)
                    drhows = pd.read_excel(pathRhows,sheet_name=sheetname + 'Std',skiprows=1,index=True)
                except:
                    print('NO rhoW values for ' + sensor + ' were computed for this campaign!')
                    continue
                if sensor == 'Hyper' and rad == 'ASD':
                    rhows  = rhows.T
                    rhows.columns = rhows.iloc[0]
                    rhows  = rhows.iloc[1:,:]
                    rhows.index.name = 'StationID'
                    drhows = drhows.T
                    drhows.columns = drhows.iloc[0]
                    drhows = drhows.iloc[1:,:]
                    drhows.index.name = 'StationID'
                else:
                    rhows  = rhows.rename(columns={'Unnamed: 0': 'StationID'})
                    rhows.set_index('StationID',inplace=True)
                    drhows = drhows.rename(columns={'Unnamed: 0': 'StationID'})
                    drhows.set_index('StationID',inplace=True)
                
                rhows.columns  = [float(c) for c in rhows.columns]
                drhows.columns = rhows.columns

                rhowB  = pd.DataFrame(index=rhows.index,columns=range(3))
                drhowB = rhowB.copy()
                for b in rhowB.columns:
                    if sensor == 'Hyper':
                        wMin  = sensors[sensor]['bands'][b][0]
                        wMax  = sensors[sensor]['bands'][b][1]
                        wBand = [r for r in rhows.columns if (r>=wMin and r<=wMax)]
                        rhowB[ b] =  rhows[wBand].mean(axis=1)
                        drhowB[b] = drhows[wBand].max( axis=1)
                    else:
                        rhowB[ b] = rhows[ sensors[sensor]['bands'][b]]
                        drhowB[b] = drhows[sensors[sensor]['bands'][b]]                        
                Algo['RAD'] = rad
                Algo.loc[np.isnan(rhowB[0]),'RAD'] = np.nan



                T  = (1/rhowB[0] + 1/rhowB[1])*rhowB[2]
                dT =   drhowB[2]*np.abs(1/rhowB[0] - 1/rhowB[1]) + \
                       drhowB[0]*rhowB[2]/(rhowB[0]**2)         + \
                       drhowB[1]*rhowB[2]/(rhowB[1]**2)                             
                Algo.loc[:,sensor] = T
                Algo.loc[:,sensor + '_Std'] = dT
                Algo.loc[:,sensor + '_CV' ] = Algo.loc[:,sensor + '_Std']/Algo.loc[:,sensor]*100

            pathXlsx = pathCampaign + '/' + campaign + '_' + algo + '.xlsx'

            wb = openpyxl.Workbook()
            wb.save(pathXlsx)

            writer = pd.ExcelWriter(pathXlsx, engine = 'openpyxl')
            writer.book = wb
            wb.remove_sheet(wb.get_sheet_by_name('Sheet'))


            # Write to Excel:
            sheetname = algo
            if sheetname in wb.sheetnames:
                wb.remove_sheet(wb.get_sheet_by_name(sheetname))
            Algo.to_excel(writer, index = True, sheet_name=sheetname,startrow = 1)
            writer.save(); writer.save()
            writer.close(); writer.close()
    return